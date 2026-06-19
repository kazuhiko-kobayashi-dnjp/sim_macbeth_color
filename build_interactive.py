#!/usr/bin/env python3
"""
Build self-contained interactive HTML for MacBeth simulation.
Usage: python build_interactive.py --excel <file.xlsx> --out macbeth_interactive.html
"""
import argparse, json
import numpy as np
import openpyxl

WL_MB = list(range(380, 731, 10))
PATCH_NAMES = [
    "1 dark skin","2 light skin","3 blue sky","4 foliage",
    "5 blue flower","6 bluish green","7 orange","8 purplish blue",
    "9 moderate red","10 purple","11 yellow green","12 orange yellow",
    "13 blue","14 green","15 red","16 yellow",
    "17 magenta","18 cyan","19 white 9.5","20 neutral 8",
    "21 neutral 6.5","22 neutral 5","23 neutral 3.5","24 black 2",
]
REFL_RAW = [
    [0.055,0.058,0.061,0.062,0.062,0.062,0.062,0.062,0.062,0.062,0.062,0.063,0.065,0.070,0.076,0.079,0.081,0.084,0.091,0.103,0.119,0.134,0.143,0.147,0.151,0.158,0.168,0.179,0.188,0.190,0.186,0.181,0.182,0.187,0.196,0.209],
    [0.117,0.143,0.175,0.191,0.196,0.199,0.204,0.213,0.228,0.251,0.280,0.309,0.329,0.333,0.315,0.286,0.273,0.276,0.277,0.289,0.339,0.420,0.488,0.525,0.546,0.562,0.578,0.595,0.612,0.625,0.638,0.656,0.678,0.700,0.717,0.734],
    [0.130,0.177,0.251,0.306,0.324,0.330,0.333,0.331,0.323,0.311,0.298,0.285,0.269,0.250,0.231,0.214,0.199,0.185,0.169,0.157,0.149,0.145,0.142,0.141,0.141,0.141,0.143,0.147,0.152,0.154,0.150,0.144,0.136,0.132,0.135,0.147],
    [0.051,0.054,0.056,0.057,0.058,0.059,0.060,0.061,0.062,0.063,0.065,0.067,0.075,0.101,0.145,0.178,0.184,0.170,0.149,0.133,0.122,0.115,0.109,0.105,0.104,0.106,0.109,0.112,0.114,0.114,0.112,0.112,0.115,0.120,0.125,0.130],
    [0.144,0.198,0.294,0.375,0.408,0.421,0.426,0.426,0.419,0.403,0.379,0.346,0.311,0.281,0.254,0.229,0.214,0.208,0.202,0.194,0.193,0.200,0.214,0.230,0.241,0.254,0.279,0.313,0.348,0.366,0.366,0.359,0.358,0.365,0.377,0.398],
    [0.136,0.179,0.247,0.297,0.320,0.337,0.355,0.381,0.419,0.466,0.510,0.546,0.567,0.574,0.569,0.551,0.524,0.488,0.445,0.400,0.350,0.299,0.252,0.221,0.204,0.196,0.191,0.188,0.191,0.199,0.212,0.223,0.232,0.233,0.229,0.229],
    [0.054,0.054,0.053,0.054,0.054,0.055,0.055,0.055,0.056,0.057,0.058,0.061,0.068,0.089,0.125,0.154,0.174,0.199,0.248,0.335,0.444,0.538,0.587,0.595,0.591,0.587,0.584,0.584,0.590,0.603,0.620,0.639,0.655,0.663,0.663,0.667],
    [0.122,0.164,0.229,0.286,0.327,0.361,0.388,0.400,0.392,0.362,0.316,0.260,0.209,0.168,0.138,0.117,0.104,0.096,0.090,0.086,0.084,0.084,0.084,0.084,0.084,0.085,0.090,0.098,0.109,0.123,0.143,0.169,0.205,0.244,0.287,0.332],
    [0.096,0.115,0.131,0.135,0.133,0.132,0.130,0.128,0.125,0.120,0.115,0.110,0.105,0.100,0.095,0.093,0.092,0.093,0.096,0.108,0.156,0.265,0.399,0.500,0.556,0.579,0.588,0.591,0.593,0.594,0.598,0.602,0.607,0.609,0.609,0.610],
    [0.092,0.116,0.146,0.169,0.178,0.173,0.158,0.139,0.119,0.101,0.087,0.075,0.066,0.060,0.056,0.053,0.051,0.051,0.052,0.052,0.051,0.052,0.058,0.073,0.096,0.119,0.141,0.166,0.194,0.227,0.265,0.309,0.355,0.396,0.436,0.478],
    [0.061,0.061,0.062,0.063,0.064,0.066,0.069,0.075,0.085,0.105,0.139,0.192,0.271,0.376,0.476,0.531,0.549,0.546,0.528,0.504,0.471,0.428,0.381,0.347,0.327,0.318,0.312,0.310,0.314,0.327,0.345,0.363,0.376,0.381,0.378,0.379],
    [0.063,0.063,0.063,0.064,0.064,0.064,0.065,0.066,0.067,0.068,0.071,0.076,0.087,0.125,0.206,0.305,0.383,0.431,0.469,0.518,0.568,0.607,0.628,0.637,0.640,0.642,0.645,0.648,0.651,0.653,0.657,0.664,0.673,0.680,0.684,0.688],
    [0.066,0.079,0.102,0.146,0.200,0.244,0.282,0.309,0.308,0.278,0.231,0.178,0.130,0.094,0.070,0.054,0.046,0.042,0.039,0.038,0.038,0.038,0.038,0.039,0.039,0.040,0.041,0.042,0.044,0.045,0.046,0.046,0.048,0.052,0.057,0.065],
    [0.052,0.053,0.054,0.055,0.057,0.059,0.061,0.066,0.075,0.093,0.125,0.178,0.246,0.307,0.337,0.334,0.317,0.293,0.262,0.230,0.198,0.165,0.135,0.115,0.104,0.098,0.094,0.092,0.093,0.097,0.102,0.108,0.113,0.115,0.114,0.114],
    [0.050,0.049,0.048,0.047,0.047,0.047,0.047,0.047,0.046,0.045,0.044,0.044,0.045,0.046,0.047,0.048,0.049,0.050,0.054,0.060,0.072,0.104,0.178,0.312,0.467,0.581,0.644,0.675,0.690,0.698,0.706,0.715,0.724,0.730,0.734,0.738],
    [0.058,0.054,0.052,0.052,0.053,0.054,0.056,0.059,0.067,0.081,0.107,0.152,0.225,0.336,0.462,0.559,0.616,0.650,0.672,0.694,0.710,0.723,0.731,0.739,0.746,0.752,0.758,0.764,0.769,0.771,0.776,0.782,0.790,0.796,0.799,0.804],
    [0.145,0.195,0.283,0.346,0.362,0.354,0.334,0.306,0.276,0.248,0.218,0.190,0.168,0.149,0.127,0.107,0.100,0.102,0.104,0.109,0.137,0.200,0.290,0.400,0.516,0.615,0.687,0.732,0.760,0.774,0.783,0.793,0.803,0.812,0.817,0.825],
    [0.108,0.141,0.192,0.236,0.261,0.286,0.317,0.353,0.390,0.426,0.446,0.444,0.423,0.385,0.337,0.283,0.231,0.185,0.146,0.118,0.101,0.090,0.082,0.076,0.074,0.073,0.073,0.074,0.076,0.077,0.076,0.075,0.073,0.072,0.074,0.079],
    [0.189,0.255,0.423,0.660,0.811,0.862,0.877,0.884,0.891,0.896,0.899,0.904,0.907,0.909,0.911,0.910,0.911,0.914,0.913,0.916,0.915,0.916,0.914,0.915,0.918,0.919,0.921,0.923,0.924,0.922,0.922,0.925,0.927,0.930,0.930,0.933],
    [0.171,0.232,0.365,0.507,0.567,0.583,0.588,0.590,0.591,0.590,0.588,0.588,0.589,0.589,0.591,0.590,0.590,0.590,0.589,0.591,0.590,0.590,0.587,0.585,0.583,0.580,0.578,0.576,0.574,0.572,0.571,0.569,0.568,0.568,0.566,0.566],
    [0.144,0.192,0.272,0.331,0.350,0.357,0.361,0.363,0.363,0.361,0.359,0.358,0.358,0.359,0.360,0.360,0.361,0.361,0.360,0.362,0.362,0.361,0.359,0.358,0.355,0.352,0.350,0.348,0.345,0.343,0.340,0.338,0.335,0.334,0.332,0.331],
    [0.105,0.131,0.163,0.180,0.186,0.190,0.193,0.194,0.194,0.192,0.191,0.191,0.191,0.192,0.192,0.192,0.192,0.192,0.192,0.193,0.192,0.192,0.191,0.189,0.188,0.186,0.184,0.182,0.181,0.179,0.178,0.176,0.174,0.173,0.172,0.171],
    [0.068,0.077,0.084,0.087,0.089,0.090,0.092,0.092,0.091,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.089,0.089,0.088,0.087,0.086,0.086,0.085,0.084,0.084,0.083,0.083,0.082,0.081,0.081,0.081],
    [0.031,0.032,0.032,0.033,0.033,0.033,0.033,0.033,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.033],
]

D65_SPD = [50.00,54.65,82.75,91.49,93.43,86.68,104.86,117.01,
           117.81,114.86,115.92,108.81,109.35,107.80,104.79,107.69,
           104.41,104.33,100.00,96.00,95.12,89.10,90.51,90.32,
           88.18,87.90,83.44,83.84,80.03,80.21,82.28,78.28,
           69.72,71.61,74.35,61.60]

def load_ccm(path):
    """Read B20:B28 from ccm_coef.xlsx.
    Order: Rr,Gr,Br, Rg,Gg,Bg, Rb,Gb,Bb
    Matrix M s.t. [R'G'B'] = [RGB] * M  (row-vector convention)
    M[i][j]: i=input(R/G/B), j=output(r/g/b)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    vals = [float(r[0]) for r in ws.iter_rows(min_row=20, max_row=28, min_col=2, max_col=2, values_only=True)]
    wb.close()
    # reshape to 3x3: row=input(R,G,B), col=output(r,g,b)
    return [[vals[0], vals[1], vals[2]],
            [vals[3], vals[4], vals[5]],
            [vals[6], vals[7], vals[8]]]

def load_optics(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["200921再構築"]
    rows = list(ws.iter_rows(min_row=68, max_row=128, values_only=True))
    wb.close()
    def col(i):
        return [float(r[i]) if r[i] is not None else 1.0 for r in rows]
    return dict(
        wl=col(1),
        cf_r=col(3),   # D: CF Red
        cf_g=col(6),   # G: CF Green (avg of Gr/Gb)
        cf_h=col(7),   # H: CF Blue
        lens=col(8),   # I: Lens
        ircf=col(9),   # J: IRCF
    )

HTML = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>MacBeth Color Simulation</title>
<script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:sans-serif;background:#1a1a2e;color:#e0e0e0;min-height:100vh}
header{background:#16213e;padding:10px 20px;display:flex;align-items:center;gap:12px;
       border-bottom:1px solid #0f3460}
header h1{font-size:16px;color:#a8d8ea;white-space:nowrap}

/* ── global controls ──────────────────────────────────── */
#gctrl{background:#16213e;padding:10px 16px;display:flex;flex-wrap:wrap;
       gap:12px;align-items:flex-end;border-bottom:2px solid #0f3460;
       position:sticky;top:0;z-index:100}
.ctrl{display:flex;flex-direction:column;gap:3px;font-size:11px}
.ctrl label{color:#a8d8ea;font-weight:bold;white-space:nowrap}
.ctrl .sub{color:#888;font-size:10px}
select{background:#0f3460;color:#e0e0e0;border:1px solid #444;
  border-radius:4px;padding:3px 6px;font-size:12px}
input[type=range]{display:block;width:110px;accent-color:#e94560;padding:0}
input[type=number]{width:60px;background:#0f3460;color:#e0e0e0;
  border:1px solid #444;border-radius:4px;padding:3px 6px;font-size:12px}
.sv{color:#e94560;font-weight:bold;min-width:26px;display:inline-block;font-size:11px}
.sep{width:1px;background:#334;align-self:stretch;margin:0 4px}
.ctrl.cf-r label{color:#e74c3c}
.ctrl.cf-g label{color:#2ecc71}
.ctrl.cf-h label{color:#3498db}

/* ── tabs ─────────────────────────────────────────────── */
.tabs{display:flex;gap:3px;padding:8px 16px 0;background:#16213e}
.tab{padding:6px 14px;border-radius:5px 5px 0 0;cursor:pointer;font-size:12px;
     background:#0f3460;color:#aaa;border:none;border-bottom:none}
.tab.active{background:#e94560;color:#fff}
.panel{display:none;padding:12px 16px}
.panel.active{display:block}

/* ── macbeth grid ─────────────────────────────────────── */
.mb-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:5px;margin-bottom:12px}
.patch{border-radius:5px;padding:8px 4px;font-size:10px;text-align:center;
       cursor:default;transition:transform .1s}
.patch:hover{transform:scale(1.06);z-index:10}
.pno{font-size:9px;opacity:.65}
.pname{font-weight:bold;font-size:10px;margin:1px 0}
.pval{font-size:9px;opacity:.85;line-height:1.3}

/* ── charts ───────────────────────────────────────────── */
.chart2{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px}

/* ── tables ───────────────────────────────────────────── */
.tbl-wrap{max-height:480px;overflow:auto;border-radius:5px}
table.tbl{width:100%;border-collapse:collapse;font-size:11px}
table.tbl th{background:#0f3460;padding:5px 7px;text-align:center;
             border:1px solid #2a3a5e;position:sticky;top:0;white-space:nowrap}
table.tbl td{padding:3px 7px;border:1px solid #222;text-align:right}
table.tbl td.left{text-align:left}
table.tbl tr:hover td{background:#1e2a40}

/* ── checklist ────────────────────────────────────────── */
.chklist{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:8px}
.chklist label{display:flex;align-items:center;gap:3px;font-size:11px;
               cursor:pointer;padding:2px 7px;border-radius:3px;border:1px solid #444}
.chklist input{accent-color:#e94560}

/* ── csv buttons ──────────────────────────────────────── */
.btn{display:inline-flex;align-items:center;gap:5px;padding:5px 12px;
     border-radius:4px;border:none;cursor:pointer;font-size:12px;font-weight:bold}
.btn-export{background:#0f3460;color:#a8d8ea;border:1px solid #3a6a9a}
.btn-export:hover{background:#1a4a7a}
.btn-import{background:#1a3a2a;color:#2ecc71;border:1px solid #2ecc71}
.btn-import:hover{background:#1e4a32}
.btn-reset{background:#3a1a1a;color:#e74c3c;border:1px solid #e74c3c}
.btn-reset:hover{background:#4a2222}
.csv-row{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:10px}
.import-box{background:#16213e;border:1px dashed #3a6a9a;border-radius:6px;
            padding:10px 14px;margin-bottom:10px;font-size:11px}
.import-box .hint{color:#888;margin-bottom:6px}
.col-map{display:flex;flex-wrap:wrap;gap:8px;margin-top:8px}
.col-map label{color:#aaa}
.col-map select{font-size:11px;padding:2px 5px}
.import-status{color:#2ecc71;font-size:11px;margin-top:6px}
.import-status.err{color:#e74c3c}

/* ── CCM panel ────────────────────────────────────────── */
.ccm-panel{background:#16213e;border:1px solid #2a4a6a;border-radius:6px;
           padding:8px 12px;margin-left:8px;font-size:11px}
.ccm-panel h3{color:#a8d8ea;font-size:11px;margin-bottom:6px}
.ccm-matrix{display:grid;grid-template-columns:repeat(4,auto);gap:2px 8px;
            font-size:11px;font-family:monospace;align-items:center}
.ccm-matrix .mhdr{color:#888;text-align:center}
.ccm-matrix .mval{text-align:right;color:#e0e0e0;min-width:52px}
.ccm-matrix .mval.pos{color:#7ec8e3}
.ccm-matrix .mval.neg{color:#e08080}
.ccm-matrix .mrow{color:#a8d8ea;font-weight:bold}
.ccm-eq{color:#888;font-size:10px;margin-top:4px}
.toggle-sw{display:flex;align-items:center;gap:6px;cursor:pointer;user-select:none}
.toggle-sw input[type=checkbox]{width:32px;height:17px;accent-color:#e94560;cursor:pointer}

/* ── normalize note ───────────────────────────────────── */
.note{font-size:11px;color:#888;margin-bottom:8px;padding:6px 10px;
      background:#16213e;border-radius:4px;border-left:3px solid #e94560}
.note b{color:#e0e0e0}
</style>
</head>
<body>
<header>
  <h1>📷 MacBeth ColorChecker Simulation</h1>
  <span style="font-size:11px;color:#666">照明光 × 分光反射率 × レンズ × IRCF × CF → R/G/B 積分値</span>
</header>

<!-- ════════════════════════════════════════════════════════
     GLOBAL CONTROLS (sticky)
═════════════════════════════════════════════════════════ -->
<div id="gctrl">
  <div class="ctrl">
    <label>照明光</label>
    <select id="g-ill" onchange="onCtrl()">
      <option value="D65" selected>D65 (6504K)</option>
      <option value="A">A光源 (2856K)</option>
    </select>
  </div>
  <div class="sep"></div>
  <div class="ctrl cf-r">
    <label>CF-R シフト <span class="sv" id="sv-cfr">0</span>nm</label>
    <input type="range" id="g-cfr" min="-15" max="15" step="1" value="0"
           oninput="document.getElementById('sv-cfr').textContent=this.value;onCtrl()">
  </div>
  <div class="ctrl cf-g">
    <label>CF-G シフト <span class="sv" id="sv-cfg">0</span>nm</label>
    <input type="range" id="g-cfg" min="-15" max="15" step="1" value="0"
           oninput="document.getElementById('sv-cfg').textContent=this.value;onCtrl()">
  </div>
  <div class="ctrl cf-h">
    <label>CF-B シフト <span class="sv" id="sv-cfh">0</span>nm</label>
    <input type="range" id="g-cfh" min="-15" max="15" step="1" value="0"
           oninput="document.getElementById('sv-cfh').textContent=this.value;onCtrl()">
  </div>
  <div class="sep"></div>
  <div class="ctrl">
    <label>IRCF シフト <span class="sv" id="sv-ircf">0</span>nm</label>
    <input type="range" id="g-ircf" min="-15" max="15" step="1" value="0"
           oninput="document.getElementById('sv-ircf').textContent=this.value;onCtrl()">
  </div>
  <div class="sep"></div>
  <div class="ctrl">
    <label>露出 <span class="sv" id="sv-ev">0</span>EV</label>
    <input type="range" id="g-ev" min="-3" max="3" step="0.25" value="0"
           oninput="document.getElementById('sv-ev').textContent=this.value;onCtrl()">
  </div>
  <div class="ctrl">
    <label>ガンマ <span class="sub">(0=なし)</span></label>
    <input type="number" id="g-gamma" value="1.5" min="0" max="4" step="0.1"
           oninput="onCtrl()">
  </div>
  <div class="ctrl">
    <label>正規化モード</label>
    <select id="g-norm" onchange="onCtrl()" title="各ch独立: 白パッチを(1,1,1)に正規化（AWB相当）。生値: Gchのみ白=1、RとBはそのまま（実際の色バランスが出る）">
      <option value="wb" selected>各ch独立 (AWB相当)</option>
      <option value="raw">生値 (G基準, WBなし)</option>
    </select>
  </div>
  <div class="ctrl">
    <label>白基準パッチ</label>
    <select id="g-white" onchange="onCtrl()">
      <option value="18" selected>19 white 9.5</option>
      <option value="17">20 neutral 8</option>
    </select>
  </div>
  <div class="sep"></div>
  <div class="ctrl">
    <label>CCM</label>
    <label class="toggle-sw">
      <input type="checkbox" id="g-ccm" onchange="onCtrl()">
      <span id="ccm-sw-label" style="color:#888">OFF</span>
    </label>
  </div>
  <div class="ccm-panel" id="ccm-panel">
    <h3>Color Correction Matrix</h3>
    <div class="ccm-matrix" id="ccm-matrix-display"></div>
    <div class="ccm-eq">[R' G' B'] = [R G B] × M　　各列和=1（白点保存）</div>
  </div>
</div>

<!-- ════════════════════════════════════════════════════════
     TABS
═════════════════════════════════════════════════════════ -->
<div class="tabs">
  <button class="tab active" onclick="showTab('tab-result',this)">🎨 カラーグリッド</button>
  <button class="tab" onclick="showTab('tab-spectra',this)">📊 分光特性グラフ</button>
  <button class="tab" onclick="showTab('tab-opttbl',this)">📋 分光特性テーブル</button>
  <button class="tab" onclick="showTab('tab-refl',this)">🌈 マクベス反射率</button>
  <button class="tab" onclick="showTab('tab-table',this)">🔢 数値テーブル</button>
</div>

<!-- ── TAB: カラーグリッド ───────────────────────────────── -->
<div id="tab-result" class="panel active">
  <div class="note" id="norm-note"></div>
  <div class="mb-grid" id="mb-grid"></div>
  <div id="plot-rgb" style="height:300px"></div>
</div>

<!-- ── TAB: 分光特性グラフ ──────────────────────────────── -->
<div id="tab-spectra" class="panel">
  <div style="display:flex;gap:12px;align-items:flex-end;margin-bottom:8px">
    <div class="ctrl">
      <label style="color:#f39c12">マクベスパッチ</label>
      <select id="sp-patch" onchange="renderSpectra()">__PATCH_OPTIONS__</select>
    </div>
  </div>
  <div class="chklist" id="sp-chklist"></div>
  <div class="chart2">
    <div id="sp-optical" style="height:340px"></div>
    <div id="sp-product" style="height:340px"></div>
  </div>
  <div id="sp-response" style="height:280px"></div>
</div>

<!-- ── TAB: 分光特性テーブル ────────────────────────────── -->
<div id="tab-opttbl" class="panel">
  <div class="csv-row">
    <button class="btn btn-export" onclick="exportOptCsv()">⬇ CSVエクスポート（現在の値）</button>
    <button class="btn btn-export" onclick="exportOptCsvRaw()">⬇ CSVエクスポート（シフトなし原点）</button>
  </div>
  <div class="import-box">
    <div class="hint">⬆ CSVインポート: 波長列 + 任意の列(cf_r / cf_g / cf_h / lens / ircf)を上書きできます。<br>
    1行目はヘッダ行。波長列ヘッダは <code>wl</code> または <code>wavelength</code> または <code>波長</code>。</div>
    <div style="display:flex;gap:8px;align-items:center">
      <input type="file" id="csv-import-file" accept=".csv" onchange="previewCsv(event)">
      <button class="btn btn-reset" onclick="resetOptData()">↺ データをリセット（元に戻す）</button>
    </div>
    <div id="csv-col-map" class="col-map" style="display:none"></div>
    <div style="margin-top:6px;display:none" id="csv-apply-row">
      <button class="btn btn-import" onclick="applyImport()">✔ この列マッピングで取り込む</button>
    </div>
    <div id="import-status" class="import-status"></div>
  </div>
  <p style="font-size:11px;color:#888;margin-bottom:8px">
    スライダーのシフト適用後の値を表示。照明光は最大値で正規化した相対値。
  </p>
  <div class="tbl-wrap">
    <table class="tbl" id="opt-table"></table>
  </div>
</div>

<!-- ── TAB: マクベス反射率 ───────────────────────────────── -->
<div id="tab-refl" class="panel">
  <div class="csv-row">
    <button class="btn btn-export" onclick="exportReflCsv()">⬇ CSVエクスポート</button>
    <button class="btn btn-reset" onclick="resetReflData()">↺ リセット（元に戻す）</button>
  </div>
  <div class="import-box">
    <div class="hint">⬆ CSVインポート: 1列目=波長(380〜730nm 10nm刻み)、2列目以降=パッチ反射率。<br>
    ヘッダ行必須。波長列は <code>wl</code> / <code>wavelength</code> / <code>波長</code>、パッチ列はパッチ名またはNo(1〜24)。</div>
    <div style="display:flex;gap:8px;align-items:center">
      <input type="file" id="refl-import-file" accept=".csv" onchange="importReflCsv(event)">
    </div>
    <div id="refl-import-status" class="import-status"></div>
  </div>
  <div class="tbl-wrap">
    <table class="tbl" id="refl-table"></table>
  </div>
</div>

<!-- ── TAB: 数値テーブル ────────────────────────────────── -->
<div id="tab-table" class="panel">
  <div class="csv-row">
    <button class="btn btn-export" onclick="exportNumCsv()">⬇ CSVエクスポート（正規化値＋8bit）</button>
    <button class="btn btn-export" onclick="exportNumCsvRaw()">⬇ CSVエクスポート（生積分値）</button>
  </div>
  <div class="note" id="tbl-norm-note"></div>
  <div class="tbl-wrap">
    <table class="tbl" id="num-table"></table>
  </div>
</div>

<script>
// ════════════════════════════════════════════════════════
//  DATA
// ════════════════════════════════════════════════════════
const DATA = __DATA_JSON__;

// ════════════════════════════════════════════════════════
//  UTILITY
// ════════════════════════════════════════════════════════
function lerp(wlSrc, valSrc, wlDst) {
  return wlDst.map(w => {
    if (w <= wlSrc[0]) return valSrc[0];
    if (w >= wlSrc[wlSrc.length-1]) return valSrc[valSrc.length-1];
    let lo = 0;
    for (let i = 0; i < wlSrc.length-1; i++) {
      if (wlSrc[i] <= w && w <= wlSrc[i+1]) { lo = i; break; }
    }
    const t = (w - wlSrc[lo])/(wlSrc[lo+1]-wlSrc[lo]);
    return valSrc[lo] + t*(valSrc[lo+1]-valSrc[lo]);
  });
}

function shiftSpec(wl, vals, shiftNm) {
  if (!shiftNm) return vals.slice();
  return lerp(wl, vals, wl.map(w => w - shiftNm));
}

function trapz(x, y) {
  let s = 0;
  for (let i = 0; i < x.length-1; i++) s += (y[i]+y[i+1])/2*(x[i+1]-x[i]);
  return s;
}

function normArr(arr) {
  const mx = Math.max(...arr);
  return mx > 0 ? arr.map(v => v/mx) : arr.slice();
}

function illA(wl) {
  const c2 = 1.4388e7, T = 2856;
  const ref = Math.exp(c2/(T*560)) - 1;
  const raw = wl.map(w => Math.pow(560/w,5)*ref/(Math.exp(c2/(T*w))-1)*100);
  return raw;
}
function illD65(wl) { return lerp(DATA.wl_mb, DATA.d65, wl); }
function getIll(wl) {
  return document.getElementById('g-ill').value === 'A' ? illA(wl) : illD65(wl);
}

function getShifts() {
  return {
    cfR:  +document.getElementById('g-cfr').value,
    cfG:  +document.getElementById('g-cfg').value,
    cfH:  +document.getElementById('g-cfh').value,
    ircf: +document.getElementById('g-ircf').value,
  };
}

function patchBgFg(Rn, Gn, Hn) {
  const r = Math.round(Math.min(1,Math.max(0,Rn))*255);
  const g = Math.round(Math.min(1,Math.max(0,Gn))*255);
  const b = Math.round(Math.min(1,Math.max(0,Hn))*255);
  const luma = 0.299*r + 0.587*g + 0.114*b;
  return {bg:`rgb(${r},${g},${b})`, fg: luma>128?'#000':'#fff'};
}

const LAYOUT = (extra) => ({
  paper_bgcolor:'#1a1a2e', plot_bgcolor:'#1a1a2e',
  font:{color:'#e0e0e0',size:11},
  xaxis:{title:'波長 (nm)',gridcolor:'#2a2a4e',tickvals:[400,450,500,550,600,650,700]},
  yaxis:{gridcolor:'#2a2a4e'},
  legend:{orientation:'h',y:1.08,font:{size:10}},
  margin:{t:36,b:44,l:52,r:10},
  ...extra,
});

// ════════════════════════════════════════════════════════
//  SIMULATION
// ════════════════════════════════════════════════════════
function simulate() {
  const wl    = DATA.wl_opt;
  const sh    = getShifts();
  const wIdx  = +document.getElementById('g-white').value;
  const norm  = document.getElementById('g-norm').value;
  const gamma = +document.getElementById('g-gamma').value;
  const ev    = +document.getElementById('g-ev').value;
  const gain  = Math.pow(2, ev);   // exposure multiplier

  const ill   = getIll(wl);
  const ircf  = shiftSpec(wl, DATA.ircf, sh.ircf);
  const cfR   = shiftSpec(wl, DATA.cf_r, sh.cfR);
  const cfG   = shiftSpec(wl, DATA.cf_g, sh.cfG);
  const cfH   = shiftSpec(wl, DATA.cf_h, sh.cfH);
  const optBase = wl.map((_,i) => DATA.lens[i] * ircf[i]);

  const raw = DATA.refl.map(refl => {
    const reflI = lerp(DATA.wl_mb, refl, wl);
    const inc   = wl.map((_,i) => ill[i] * reflI[i] * optBase[i]);
    return {
      R: trapz(wl, inc.map((v,i) => v*cfR[i])),
      G: trapz(wl, inc.map((v,i) => v*cfG[i])),
      H: trapz(wl, inc.map((v,i) => v*cfH[i])),
    };
  });

  const wRef = raw[wIdx];
  const scaleR = norm === 'wb' ? wRef.R : wRef.G;
  const scaleG = wRef.G;
  const scaleH = norm === 'wb' ? wRef.H : wRef.G;

  const useCcm = document.getElementById('g-ccm').checked;
  const M = DATA.ccm;  // M[i][j]: input i → output j

  return raw.map(r => {
    // linear normalized value × exposure gain
    let Rn = r.R / scaleR * gain;
    let Gn = r.G / scaleG * gain;
    let Hn = r.H / scaleH * gain;

    // CCM: [R'G'B'] = [Rn Gn Hn] * M
    if (useCcm) {
      const Rp = Rn*M[0][0] + Gn*M[1][0] + Hn*M[2][0];
      const Gp = Rn*M[0][1] + Gn*M[1][1] + Hn*M[2][1];
      const Hp = Rn*M[0][2] + Gn*M[1][2] + Hn*M[2][2];
      Rn = Rp; Gn = Gp; Hn = Hp;
    }

    // gamma-corrected 8-bit (clipped to [0,1] before power)
    const applyGamma = (v) => {
      if (gamma <= 0) return Math.min(1, Math.max(0, v));
      return Math.pow(Math.min(1, Math.max(0, v)), 1/gamma);
    };
    const Rg = applyGamma(Rn);
    const Gg = applyGamma(Gn);
    const Hg = applyGamma(Hn);
    const R8 = Math.round(Rg * 255);
    const G8 = Math.round(Gg * 255);
    const H8 = Math.round(Hg * 255);

    return {R_raw:r.R, G_raw:r.G, H_raw:r.H,
            Rn, Gn, Hn,           // linear after CCM (for charts)
            Rg, Gg, Hg,           // gamma-corrected [0-1] (for patch color)
            R8, G8, H8};          // 8-bit
  });
}

// ════════════════════════════════════════════════════════
//  TAB: カラーグリッド
// ════════════════════════════════════════════════════════
function renderResult() {
  const res   = simulate();
  const gamma = +document.getElementById('g-gamma').value;
  const norm  = document.getElementById('g-norm').value;

  // note
  const noteEl = document.getElementById('norm-note');
  if (norm === 'wb') {
    noteEl.innerHTML = '<b>各ch独立正規化（AWB相当）:</b> R・G・Bそれぞれを白パッチ値で割るため、白は常に(1,1,1)になります。CFシフトの色バランス変化は見えません。';
  } else {
    noteEl.innerHTML = '<b>生値（G基準・WBなし）:</b> 全chをG_whiteで割る共通スケール。白パッチのRとBが1になるとは限らず、CFシフトの色バランス変化が見えます。';
  }

  // grid
  const grid = document.getElementById('mb-grid');
  grid.innerHTML = '';
  res.forEach((r,i) => {
    const {bg,fg} = patchBgFg(r.Rg, r.Gg, r.Hg);  // use gamma-corrected values for display
    const name = DATA.patch_names[i].replace(/^\d+\s/,'');
    const v = `R=${r.R8} G=${r.G8} B=${r.H8}`;
    const d = document.createElement('div');
    d.className = 'patch';
    d.style.cssText = `background:${bg};color:${fg}`;
    d.innerHTML = `<div class="pno">${i+1}</div><div class="pname">${name}</div><div class="pval">${v}</div>`;
    grid.appendChild(d);
  });

  // bar chart
  const names = DATA.patch_names.map(n => n.replace(/^\d+\s/,''));
  Plotly.react('plot-rgb', [
    {x:names, y:res.map(r=>+r.Rn.toFixed(4)), name:'R', type:'bar', marker:{color:'#e74c3c'}},
    {x:names, y:res.map(r=>+r.Gn.toFixed(4)), name:'G', type:'bar', marker:{color:'#2ecc71'}},
    {x:names, y:res.map(r=>+r.Hn.toFixed(4)), name:'B', type:'bar', marker:{color:'#3498db'}},
  ], LAYOUT({barmode:'group', height:300,
    xaxis:{...LAYOUT({}).xaxis, tickangle:-45, title:null},
    yaxis:{...LAYOUT({}).yaxis, title:'正規化値'},
    margin:{t:20,b:110,l:50,r:10}}));
}

// ════════════════════════════════════════════════════════
//  TAB: 分光特性グラフ
// ════════════════════════════════════════════════════════
const SP_CURVES = [
  {id:'sp_ill',  label:'照明光',    color:'#f0e68c'},
  {id:'sp_lens', label:'レンズ',    color:'#b0c4de'},
  {id:'sp_ircf', label:'IRCF',      color:'#dda0dd'},
  {id:'sp_cfr',  label:'CF-R',      color:'#e74c3c'},
  {id:'sp_cfg',  label:'CF-G',      color:'#2ecc71'},
  {id:'sp_cfh',  label:'CF-B',      color:'#3498db'},
  {id:'sp_refl', label:'分光反射率', color:'#f39c12'},
];

function buildSpChecklist() {
  const cl = document.getElementById('sp-chklist');
  SP_CURVES.forEach(c => {
    const lbl = document.createElement('label');
    lbl.style.borderColor = c.color;
    lbl.innerHTML = `<input type="checkbox" id="${c.id}" checked onchange="renderSpectra()">
      <span style="color:${c.color}">${c.label}</span>`;
    cl.appendChild(lbl);
  });
}
function chk(id){const e=document.getElementById(id);return e&&e.checked;}

function renderSpectra() {
  const wl  = DATA.wl_opt;
  const sh  = getShifts();
  const pi  = +document.getElementById('sp-patch').value;

  const ill   = normArr(getIll(wl));
  const ircf  = shiftSpec(wl, DATA.ircf, sh.ircf);
  const cfR   = shiftSpec(wl, DATA.cf_r, sh.cfR);
  const cfG   = shiftSpec(wl, DATA.cf_g, sh.cfG);
  const cfH   = shiftSpec(wl, DATA.cf_h, sh.cfH);
  const refl  = lerp(DATA.wl_mb, DATA.refl[pi], wl);
  const opt   = wl.map((_,i) => DATA.lens[i]*ircf[i]);
  const inc   = wl.map((_,i) => ill[i]*refl[i]*opt[i]);

  // left: individual spectra
  const t1 = [];
  if(chk('sp_ill'))  t1.push({x:wl,y:ill,   name:'照明光(norm)',line:{color:'#f0e68c',width:2}});
  if(chk('sp_refl')) t1.push({x:wl,y:refl,  name:'分光反射率', line:{color:'#f39c12',width:2}});
  if(chk('sp_lens')) t1.push({x:wl,y:DATA.lens,name:'レンズ', line:{color:'#b0c4de',width:2}});
  if(chk('sp_ircf')) t1.push({x:wl,y:ircf,  name:'IRCF',      line:{color:'#dda0dd',width:2}});
  if(chk('sp_cfr'))  t1.push({x:wl,y:cfR,   name:'CF-R',      line:{color:'#e74c3c',width:2}});
  if(chk('sp_cfg'))  t1.push({x:wl,y:cfG,   name:'CF-G',      line:{color:'#2ecc71',width:2}});
  if(chk('sp_cfh'))  t1.push({x:wl,y:cfH,   name:'CF-B',      line:{color:'#3498db',width:2}});
  Plotly.react('sp-optical', t1, LAYOUT({height:340,
    yaxis:{...LAYOUT({}).yaxis,title:'透過率 / 反射率 (相対)'},
    title:{text:'各特性（個別）',font:{color:'#a8d8ea',size:12}}}));

  // right: stacked optical path
  const t2 = [
    {x:wl,y:opt, name:'レンズ×IRCF',       line:{color:'#dda0dd',width:2}},
    {x:wl,y:refl,name:'分光反射率',          line:{color:'#f39c12',width:2,dash:'dot'}},
    {x:wl,y:inc, name:'照明×反射×光学系',    fill:'tozeroy',
     line:{color:'#f0e68c',width:1.5},fillcolor:'rgba(240,230,140,0.12)'},
  ];
  Plotly.react('sp-product', t2, LAYOUT({height:340,
    yaxis:{...LAYOUT({}).yaxis,title:'相対強度'},
    title:{text:'光路積（掛け合わせ）',font:{color:'#a8d8ea',size:12}}}));

  // bottom: R/G/B response (incident × CF)
  const t3 = [
    {x:wl,y:inc.map((v,i)=>v*cfR[i]),name:'R応答',fill:'tozeroy',
     line:{color:'#e74c3c',width:1.5},fillcolor:'rgba(231,76,60,0.22)'},
    {x:wl,y:inc.map((v,i)=>v*cfG[i]),name:'G応答',fill:'tozeroy',
     line:{color:'#2ecc71',width:1.5},fillcolor:'rgba(46,204,113,0.22)'},
    {x:wl,y:inc.map((v,i)=>v*cfH[i]),name:'B応答',fill:'tozeroy',
     line:{color:'#3498db',width:1.5},fillcolor:'rgba(52,152,219,0.22)'},
  ];
  Plotly.react('sp-response', t3, LAYOUT({height:280,
    yaxis:{...LAYOUT({}).yaxis,title:'入射光 × CF（この面積がR/G/B積分値）'},
    title:{text:'R/G/B 分光応答',font:{color:'#a8d8ea',size:12}}}));
}

// ════════════════════════════════════════════════════════
//  TAB: 分光特性テーブル
// ════════════════════════════════════════════════════════
function renderOptTable() {
  const wl   = DATA.wl_opt;
  const sh   = getShifts();
  const illN = normArr(getIll(wl));
  const ircf = shiftSpec(wl, DATA.ircf, sh.ircf);
  const cfR  = shiftSpec(wl, DATA.cf_r, sh.cfR);
  const cfG  = shiftSpec(wl, DATA.cf_g, sh.cfG);
  const cfH  = shiftSpec(wl, DATA.cf_h, sh.cfH);

  const sfx = s => s >= 0 ? `+${s}` : `${s}`;
  const hdrs = [
    '波長(nm)',
    `照明光<br>(${document.getElementById('g-ill').value}, norm)`,
    'レンズ',
    `IRCF<br>(${sfx(sh.ircf)}nm)`,
    `CF-R<br>(${sfx(sh.cfR)}nm)`,
    `CF-G<br>(${sfx(sh.cfG)}nm)`,
    `CF-B<br>(${sfx(sh.cfH)}nm)`,
    'レンズ×IRCF',
  ];

  let html = '<thead><tr>' + hdrs.map(h=>`<th>${h}</th>`).join('') + '</tr></thead><tbody>';
  for (let i=0; i<wl.length; i++) {
    const opt = DATA.lens[i]*ircf[i];
    html += `<tr>
      <td style="color:#a8d8ea;font-weight:bold">${wl[i].toFixed(0)}</td>
      <td>${illN[i].toFixed(4)}</td>
      <td>${DATA.lens[i].toFixed(4)}</td>
      <td>${ircf[i].toFixed(4)}</td>
      <td style="color:#e74c3c">${cfR[i].toFixed(4)}</td>
      <td style="color:#2ecc71">${cfG[i].toFixed(4)}</td>
      <td style="color:#3498db">${cfH[i].toFixed(4)}</td>
      <td>${opt.toFixed(4)}</td>
    </tr>`;
  }
  html += '</tbody>';
  document.getElementById('opt-table').innerHTML = html;
}

// ════════════════════════════════════════════════════════
//  TAB: 数値テーブル
// ════════════════════════════════════════════════════════
function renderNumTable() {
  const res  = simulate();
  const norm = document.getElementById('g-norm').value;
  const noteEl = document.getElementById('tbl-norm-note');
  noteEl.innerHTML = norm === 'wb'
    ? '<b>各ch独立正規化:</b> white patch のR/G/Bそれぞれで割算。白は常に(1,1,1)。'
    : '<b>生値 (G基準):</b> G_whiteのみで全ch割算。色バランスが正直に出ます。';

  const hdrs = ['No','パッチ名','R_raw','G_raw','B_raw',
                'R_norm(lin)','G_norm(lin)','B_norm(lin)','R8','G8','B8'];
  let html = '<thead><tr>'+hdrs.map(h=>`<th>${h}</th>`).join('')+'</tr></thead><tbody>';
  res.forEach((r,i) => {
    const {bg,fg} = patchBgFg(r.Rg, r.Gg, r.Hg);
    html += `<tr>
      <td>${i+1}</td>
      <td class="left" style="background:${bg};color:${fg}">${DATA.patch_names[i]}</td>
      <td>${r.R_raw.toFixed(3)}</td><td>${r.G_raw.toFixed(3)}</td><td>${r.H_raw.toFixed(3)}</td>
      <td>${r.Rn.toFixed(4)}</td><td>${r.Gn.toFixed(4)}</td><td>${r.Hn.toFixed(4)}</td>
      <td>${r.R8}</td><td>${r.G8}</td><td>${r.H8}</td>
    </tr>`;
  });
  html += '</tbody>';
  document.getElementById('num-table').innerHTML = html;
}

// ════════════════════════════════════════════════════════
//  CSV EXPORT
// ════════════════════════════════════════════════════════
function downloadCsv(filename, rows) {
  const bom = '﻿';  // UTF-8 BOM for Excel
  const text = bom + rows.map(r => r.join(',')).join('\r\n');
  const a = document.createElement('a');
  a.href = URL.createObjectURL(new Blob([text], {type:'text/csv;charset=utf-8'}));
  a.download = filename;
  a.click();
}

function exportOptCsv() {
  const wl   = DATA.wl_opt;
  const sh   = getShifts();
  const ill  = document.getElementById('g-ill').value;
  const illN = normArr(getIll(wl));
  const ircf = shiftSpec(wl, DATA.ircf, sh.ircf);
  const cfR  = shiftSpec(wl, DATA.cf_r, sh.cfR);
  const cfG  = shiftSpec(wl, DATA.cf_g, sh.cfG);
  const cfH  = shiftSpec(wl, DATA.cf_h, sh.cfH);
  const rows = [['wl', `ill_${ill}_norm`, 'lens', 'ircf', 'cf_r', 'cf_g', 'cf_h', 'lens_x_ircf']];
  for (let i=0; i<wl.length; i++) {
    rows.push([wl[i], illN[i].toFixed(5), DATA.lens[i].toFixed(5),
               ircf[i].toFixed(5), cfR[i].toFixed(5), cfG[i].toFixed(5),
               cfH[i].toFixed(5), (DATA.lens[i]*ircf[i]).toFixed(5)]);
  }
  downloadCsv('optical_spectra_shifted.csv', rows);
}

function exportOptCsvRaw() {
  const wl  = DATA.wl_opt;
  const rows = [['wl', 'lens', 'ircf', 'cf_r', 'cf_g', 'cf_h']];
  for (let i=0; i<wl.length; i++) {
    rows.push([wl[i], DATA.lens[i].toFixed(5), DATA.ircf[i].toFixed(5),
               DATA.cf_r[i].toFixed(5), DATA.cf_g[i].toFixed(5), DATA.cf_h[i].toFixed(5)]);
  }
  downloadCsv('optical_spectra_raw.csv', rows);
}

function exportNumCsv() {
  const res  = simulate();
  const rows = [['No','patch_name','R_raw','G_raw','B_raw',
                 'R_norm_lin','G_norm_lin','B_norm_lin','R8','G8','B8']];
  res.forEach((r,i) => {
    rows.push([i+1, DATA.patch_names[i],
               r.R_raw.toFixed(4), r.G_raw.toFixed(4), r.H_raw.toFixed(4),
               r.Rn.toFixed(5), r.Gn.toFixed(5), r.Hn.toFixed(5),
               r.R8, r.G8, r.H8]);
  });
  downloadCsv('macbeth_result.csv', rows);
}

function exportNumCsvRaw() {
  const res  = simulate();
  const rows = [['No','patch_name','R_raw','G_raw','B_raw']];
  res.forEach((r,i) => {
    rows.push([i+1, DATA.patch_names[i],
               r.R_raw.toFixed(4), r.G_raw.toFixed(4), r.H_raw.toFixed(4)]);
  });
  downloadCsv('macbeth_result_raw.csv', rows);
}

// ════════════════════════════════════════════════════════
//  TAB: マクベス反射率テーブル
// ════════════════════════════════════════════════════════
function renderReflTable() {
  const wl = DATA.wl_mb;
  const hdrs = ['波長(nm)', ...DATA.patch_names.map(n => n.replace(/^\d+\s/,''))];
  let html = '<thead><tr>' + hdrs.map(h=>`<th style="white-space:nowrap">${h}</th>`).join('') + '</tr></thead><tbody>';
  for (let wi=0; wi<wl.length; wi++) {
    html += `<tr><td style="color:#a8d8ea;font-weight:bold">${wl[wi]}</td>`;
    for (let pi=0; pi<DATA.refl.length; pi++) {
      html += `<td>${DATA.refl[pi][wi].toFixed(4)}</td>`;
    }
    html += '</tr>';
  }
  html += '</tbody>';
  document.getElementById('refl-table').innerHTML = html;
}

function exportReflCsv() {
  const wl = DATA.wl_mb;
  const rows = [['wl', ...DATA.patch_names]];
  for (let wi=0; wi<wl.length; wi++) {
    rows.push([wl[wi], ...DATA.refl.map(r => r[wi].toFixed(4))]);
  }
  downloadCsv('macbeth_reflectance.csv', rows);
}

function resetReflData() {
  for (let pi=0; pi<DATA.refl.length; pi++) {
    DATA.refl[pi] = DATA_ORIG.refl[pi].slice();
  }
  document.getElementById('refl-import-status').className = 'import-status';
  document.getElementById('refl-import-status').textContent = '元データに戻しました。';
  document.getElementById('refl-import-file').value = '';
  renderActive();
}

function importReflCsv(e) {
  const file = e.target.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = ev => {
    const parsed = parseCsvText(ev.target.result);
    if (!parsed) { showReflStatus('CSVの解析に失敗しました。', true); return; }
    const {headers, data} = parsed;
    const wlCol = findWlCol(headers);
    const csvWl = data.map(r => r[wlCol]);

    // match each CSV column to a patch by name or number
    let applied = 0;
    for (let ci=0; ci<headers.length; ci++) {
      if (ci === wlCol) continue;
      const h = headers[ci].trim();
      // try numeric match (patch No)
      const num = parseInt(h);
      let pi = -1;
      if (!isNaN(num) && num >= 1 && num <= 24) {
        pi = num - 1;
      } else {
        // fuzzy name match
        pi = DATA.patch_names.findIndex(n =>
          n.toLowerCase().includes(h.toLowerCase()) ||
          h.toLowerCase().includes(n.replace(/^\d+\s/,'').toLowerCase())
        );
      }
      if (pi < 0) continue;
      const vals = data.map(r => isNaN(r[ci]) ? 0 : r[ci]);
      DATA.refl[pi] = lerp(csvWl, vals, DATA.wl_mb);
      applied++;
    }
    if (applied === 0) { showReflStatus('マッチするパッチ列が見つかりませんでした。', true); return; }
    showReflStatus(`✔ ${applied}パッチの反射率を更新しました。`);
    renderActive();
  };
  reader.readAsText(file, 'utf-8');
}

function showReflStatus(msg, isErr=false) {
  const el = document.getElementById('refl-import-status');
  el.className = 'import-status' + (isErr ? ' err' : '');
  el.textContent = msg;
}

// ════════════════════════════════════════════════════════
//  CSV IMPORT
// ════════════════════════════════════════════════════════
// Original data backup (set once at init)
const DATA_ORIG = {};
let csvParsed = null;  // {headers, wl, cols}

function initOrig() {
  DATA_ORIG.wl_opt = DATA.wl_opt.slice();
  DATA_ORIG.cf_r   = DATA.cf_r.slice();
  DATA_ORIG.cf_g   = DATA.cf_g.slice();
  DATA_ORIG.cf_h   = DATA.cf_h.slice();
  DATA_ORIG.lens   = DATA.lens.slice();
  DATA_ORIG.ircf   = DATA.ircf.slice();
  DATA_ORIG.refl   = DATA.refl.map(r => r.slice());
}

function resetOptData() {
  Object.assign(DATA, DATA_ORIG);
  DATA.cf_r  = DATA_ORIG.cf_r.slice();
  DATA.cf_g  = DATA_ORIG.cf_g.slice();
  DATA.cf_h  = DATA_ORIG.cf_h.slice();
  DATA.lens  = DATA_ORIG.lens.slice();
  DATA.ircf  = DATA_ORIG.ircf.slice();
  document.getElementById('import-status').className = 'import-status';
  document.getElementById('import-status').textContent = '元データに戻しました。';
  document.getElementById('csv-col-map').style.display = 'none';
  document.getElementById('csv-apply-row').style.display = 'none';
  document.getElementById('csv-import-file').value = '';
  csvParsed = null;
  renderActive();
}

function parseCsvText(text) {
  const lines = text.replace(/\r\n/g,'\n').replace(/\r/g,'\n').split('\n')
    .map(l=>l.trim()).filter(l=>l);
  if (lines.length < 2) return null;
  const headers = lines[0].split(',').map(h=>h.trim().replace(/^"|"$/g,''));
  const data = lines.slice(1).map(l => l.split(',').map(v=>parseFloat(v.trim().replace(/^"|"$/g,''))));
  return {headers, data};
}

function findWlCol(headers) {
  const wlNames = ['wl','wavelength','波長','nm','lambda'];
  for (let i=0; i<headers.length; i++) {
    if (wlNames.includes(headers[i].toLowerCase())) return i;
  }
  return 0;  // fallback: first column
}

function previewCsv(e) {
  const file = e.target.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = ev => {
    const parsed = parseCsvText(ev.target.result);
    if (!parsed) {
      showImportStatus('CSVの解析に失敗しました。', true); return;
    }
    csvParsed = parsed;
    buildColMap(parsed.headers);
  };
  reader.readAsText(file, 'utf-8');
}

const TARGET_COLS = [
  {key:'cf_r',  label:'CF-R',  color:'#e74c3c'},
  {key:'cf_g',  label:'CF-G',  color:'#2ecc71'},
  {key:'cf_h',  label:'CF-B',  color:'#3498db'},
  {key:'lens',  label:'レンズ', color:'#b0c4de'},
  {key:'ircf',  label:'IRCF',  color:'#dda0dd'},
];
const COL_ALIASES = {
  cf_r: ['cf_r','cfr','r','red','cf-r'],
  cf_g: ['cf_g','cfg','g','green','cf-g'],
  cf_h: ['cf_h','cfh','b','blue','cf-b','cf_b'],
  lens: ['lens','lens_t'],
  ircf: ['ircf','ircf_t'],
};

function guessCol(headers, key) {
  const aliases = COL_ALIASES[key] || [key];
  for (let i=0; i<headers.length; i++) {
    if (aliases.includes(headers[i].toLowerCase())) return i;
  }
  return -1;
}

function buildColMap(headers) {
  const mapDiv = document.getElementById('csv-col-map');
  const applyRow = document.getElementById('csv-apply-row');
  const opts = ['（使わない）', ...headers].map((h,i) =>
    `<option value="${i-1}">${h}</option>`).join('');

  mapDiv.innerHTML = TARGET_COLS.map(t => {
    const guess = guessCol(headers, t.key);
    const sel = ['（使わない）', ...headers].map((h,i) => {
      const val = i - 1;
      const sel = val === guess ? ' selected' : '';
      return `<option value="${val}"${sel}>${h}</option>`;
    }).join('');
    return `<label style="color:${t.color}">${t.label}:
      <select id="map-${t.key}" style="background:#0f3460;color:#e0e0e0;
        border:1px solid #444;border-radius:3px;padding:2px 4px;font-size:11px">${sel}</select>
    </label>`;
  }).join('');

  mapDiv.style.display = 'flex';
  applyRow.style.display = 'block';
  showImportStatus(`${headers.length}列 × ${csvParsed.data.length}行 を読み込みました。列マッピングを確認して「取り込む」を押してください。`);
}

function applyImport() {
  if (!csvParsed) return;
  const {headers, data} = csvParsed;
  const wlCol = findWlCol(headers);
  const csvWl = data.map(r => r[wlCol]).filter(v => !isNaN(v));
  if (csvWl.length === 0) { showImportStatus('波長列が見つかりません。', true); return; }

  let applied = [];
  TARGET_COLS.forEach(t => {
    const colIdx = +document.getElementById(`map-${t.key}`).value;
    if (colIdx < 0) return;
    const vals = data.map(r => isNaN(r[colIdx]) ? 1.0 : r[colIdx]);
    // interpolate to DATA.wl_opt grid
    DATA[t.key] = lerp(csvWl, vals, DATA.wl_opt);
    applied.push(t.label);
  });

  if (applied.length === 0) { showImportStatus('取り込む列が1つも選択されていません。', true); return; }
  showImportStatus(`✔ ${applied.join(', ')} を更新しました（波長: ${csvWl[0]}〜${csvWl[csvWl.length-1]}nm, ${csvWl.length}点 → ${DATA.wl_opt.length}点に補間）`);
  renderActive();
}

function showImportStatus(msg, isErr=false) {
  const el = document.getElementById('import-status');
  el.className = 'import-status' + (isErr ? ' err' : '');
  el.textContent = msg;
}

// ════════════════════════════════════════════════════════
//  ROUTING
// ════════════════════════════════════════════════════════
let activeTab = 'tab-result';

function showTab(id, btn) {
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
  activeTab = id;
  renderActive();
}

function renderActive() {
  if (activeTab==='tab-result')  renderResult();
  if (activeTab==='tab-spectra') renderSpectra();
  if (activeTab==='tab-opttbl')  renderOptTable();
  if (activeTab==='tab-refl')    renderReflTable();
  if (activeTab==='tab-table')   renderNumTable();
}

function onCtrl() {
  // update CCM label
  const on = document.getElementById('g-ccm').checked;
  document.getElementById('ccm-sw-label').textContent = on ? 'ON' : 'OFF';
  document.getElementById('ccm-sw-label').style.color = on ? '#e94560' : '#888';
  renderActive();
}

function renderCcmMatrix() {
  const M = DATA.ccm;
  const labels = ['R','G','B'];
  const outLabels = ["R'","G'","B'"];
  let html = '';
  // header row
  html += `<div class="mhdr"></div>`;
  outLabels.forEach(l => html += `<div class="mhdr">${l}</div>`);
  // data rows
  for (let i=0; i<3; i++) {
    html += `<div class="mrow">${labels[i]}</div>`;
    for (let j=0; j<3; j++) {
      const v = M[i][j];
      const cls = v < 0 ? 'neg' : 'pos';
      html += `<div class="mval ${cls}">${v.toFixed(4)}</div>`;
    }
  }
  document.getElementById('ccm-matrix-display').innerHTML = html;
}

// ════════════════════════════════════════════════════════
//  INIT
// ════════════════════════════════════════════════════════
initOrig();
buildSpChecklist();
renderCcmMatrix();
renderResult();
</script>
</body>
</html>
"""

def build_patch_options():
    return '\n'.join(f'<option value="{i}">{n}</option>' for i,n in enumerate(PATCH_NAMES))

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True)
    p.add_argument("--ccm", default="ccm_coef.xlsx")
    p.add_argument("--out", default="macbeth_interactive.html")
    args = p.parse_args()

    print(f"Loading {args.excel} ...")
    opt = load_optics(args.excel)
    print(f"Loading {args.ccm} ...")
    ccm = load_ccm(args.ccm)
    def r5(lst): return [round(float(v), 5) for v in lst]

    data = dict(
        wl_mb      = WL_MB,
        wl_opt     = [float(w) for w in opt["wl"]],
        d65        = D65_SPD,
        cf_r       = r5(opt["cf_r"]),
        cf_g       = r5(opt["cf_g"]),
        cf_h       = r5(opt["cf_h"]),
        lens       = r5(opt["lens"]),
        ircf       = r5(opt["ircf"]),
        refl       = [[round(v,4) for v in row] for row in REFL_RAW],
        patch_names= PATCH_NAMES,
        ccm        = ccm,
    )
    html = HTML.replace("__DATA_JSON__", json.dumps(data))
    html = html.replace("__PATCH_OPTIONS__", build_patch_options())
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Saved: {args.out}")

if __name__ == "__main__":
    main()
