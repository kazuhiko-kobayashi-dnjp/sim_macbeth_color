#!/usr/bin/env python3
"""
Batch export: 8 CF-shift patterns × all data → Excel
CFシフト: R/G/B 各 0 or +10nm の 2^3=8パターン (D65固定)

Output sheets per pattern:
  - RGB積分値・正規化・8bit (CCMあり/なし)
  - 分光特性 (lens, ircf, cf_r/g/h shifted)
  - マクベス反射率

Summary sheet: 全パターンのRGB8bit値を横並び

Usage:
  python batch_export.py \
    --excel 201006_車両WSカメラ影響定量化まとめ２.xlsx \
    --ccm ccm_coef.xlsx \
    --out batch_result.xlsx
"""
import argparse
import itertools
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── shared data ───────────────────────────────────────────────────────────────
WL_MB = np.arange(380, 731, 10)
PATCH_NAMES = [
    "1 dark skin","2 light skin","3 blue sky","4 foliage",
    "5 blue flower","6 bluish green","7 orange","8 purplish blue",
    "9 moderate red","10 purple","11 yellow green","12 orange yellow",
    "13 blue","14 green","15 red","16 yellow",
    "17 magenta","18 cyan","19 white 9.5","20 neutral 8",
    "21 neutral 6.5","22 neutral 5","23 neutral 3.5","24 black 2",
]
REFL = np.array([
    [0.055,0.058,0.061,0.062,0.062,0.062,0.062,0.062,0.062,0.062,0.062,0.063,0.065,0.070,0.076,0.079,0.081,0.084,0.091,0.103,0.119,0.134,0.143,0.147,0.151,0.158,0.168,0.179,0.188,0.190,0.186,0.181,0.182,0.187,0.196,0.209],
    [0.117,0.143,0.175,0.191,0.196,0.199,0.204,0.213,0.228,0.251,0.280,0.309,0.329,0.333,0.315,0.286,0.273,0.276,0.277,0.289,0.339,0.420,0.488,0.525,0.546,0.562,0.578,0.595,0.612,0.625,0.638,0.656,0.678,0.700,0.717,0.734],
    [0.130,0.177,0.251,0.306,0.324,0.330,0.333,0.331,0.323,0.311,0.298,0.285,0.269,0.250,0.231,0.214,0.199,0.185,0.169,0.157,0.149,0.145,0.142,0.141,0.141,0.141,0.143,0.147,0.152,0.154,0.150,0.144,0.136,0.132,0.135,0.147],
    [0.051,0.054,0.056,0.057,0.058,0.059,0.060,0.061,0.062,0.063,0.065,0.067,0.075,0.101,0.145,0.178,0.184,0.170,0.149,0.133,0.122,0.115,0.109,0.105,0.104,0.106,0.109,0.112,0.114,0.114,0.112,0.112,0.115,0.120,0.125,0.130],
    [0.144,0.198,0.294,0.375,0.408,0.421,0.426,0.426,0.419,0.403,0.379,0.346,0.311,0.281,0.254,0.229,0.214,0.208,0.202,0.194,0.193,0.200,0.214,0.230,0.241,0.254,0.279,0.313,0.348,0.366,0.366,0.359,0.358,0.365,0.377,0.398],
    [0.136,0.179,0.247,0.297,0.320,0.337,0.355,0.381,0.419,0.466,0.510,0.546,0.567,0.574,0.569,0.551,0.524,0.488,0.445,0.400,0.350,0.299,0.252,0.221,0.204,0.196,0.191,0.188,0.191,0.199,0.212,0.223,0.232,0.233,0.229,0.229],
    [0.054,0.054,0.053,0.054,0.054,0.055,0.055,0.055,0.056,0.057,0.058,0.061,0.068,0.089,0.125,0.154,0.174,0.199,0.248,0.335,0.444,0.538,0.587,0.595,0.591,0.587,0.584,0.584,0.590,0.603,0.620,0.639,0.655,0.663,0.663,0.667],
    [0.122,0.164,0.229,0.286,0.327,0.361,0.388,0.400,0.392,0.362,0.316,0.260,0.209,0.168,0.138,0.117,0.104,0.096,0.090,0.086,0.084,0.084,0.084,0.084,0.084,0.085,0.090,0.098,0.109,0.123,0.143,0.169,0.205,0.244,0.287,0.332],
    [0.096,0.115,0.131,0.135,0.133,0.132,0.130,0.128,0.125,0.120,0.115,0.110,0.105,0.100,0.095,0.093,0.092,0.093,0.096,0.108,0.156,0.265,0.399,0.500,0.556,0.579,0.588,0.591,0.593,0.594,0.598,0.602,0.607,0.609,0.609,0.610],
    [0.092,0.116,0.146,0.169,0.178,0.173,0.158,0.139,0.119,0.101,0.087,0.075,0.066,0.060,0.056,0.053,0.051,0.051,0.052,0.052,0.051,0.052,0.058,0.073,0.096,0.119,0.141,0.166,0.194,0.227,0.265,0.309,0.355,0.396,0.436,0.478],
    [0.061,0.061,0.062,0.063,0.064,0.066,0.069,0.075,0.085,0.105,0.139,0.192,0.271,0.376,0.476,0.531,0.549,0.546,0.528,0.504,0.471,0.428,0.381,0.347,0.327,0.318,0.312,0.310,0.314,0.327,0.345,0.363,0.376,0.381,0.378,0.379],
    [0.063,0.063,0.063,0.064,0.064,0.064,0.065,0.066,0.067,0.068,0.071,0.076,0.087,0.125,0.206,0.305,0.383,0.431,0.469,0.518,0.568,0.607,0.628,0.637,0.640,0.642,0.645,0.648,0.651,0.653,0.657,0.664,0.673,0.680,0.684,0.688],
    [0.066,0.079,0.102,0.146,0.200,0.244,0.282,0.309,0.308,0.278,0.231,0.178,0.130,0.094,0.070,0.054,0.046,0.042,0.039,0.038,0.038,0.038,0.038,0.039,0.039,0.040,0.041,0.042,0.044,0.045,0.046,0.046,0.048,0.052,0.057,0.065],
    [0.052,0.053,0.054,0.055,0.057,0.059,0.061,0.066,0.075,0.093,0.125,0.178,0.246,0.307,0.337,0.334,0.317,0.293,0.262,0.230,0.198,0.165,0.135,0.115,0.104,0.098,0.094,0.092,0.093,0.097,0.102,0.108,0.113,0.115,0.114,0.114],
    [0.050,0.049,0.048,0.047,0.047,0.047,0.047,0.047,0.046,0.045,0.044,0.044,0.045,0.046,0.047,0.048,0.049,0.050,0.054,0.060,0.072,0.104,0.178,0.312,0.467,0.581,0.644,0.675,0.690,0.698,0.706,0.715,0.724,0.730,0.734,0.738],
    [0.058,0.054,0.052,0.052,0.053,0.054,0.056,0.059,0.067,0.081,0.107,0.152,0.225,0.336,0.462,0.559,0.616,0.650,0.672,0.694,0.710,0.723,0.731,0.739,0.746,0.752,0.758,0.764,0.769,0.771,0.776,0.782,0.790,0.796,0.799,0.804],
    [0.145,0.195,0.283,0.346,0.362,0.354,0.334,0.306,0.276,0.248,0.218,0.190,0.168,0.149,0.127,0.107,0.100,0.102,0.104,0.109,0.137,0.200,0.290,0.400,0.516,0.615,0.687,0.732,0.760,0.774,0.783,0.793,0.803,0.812,0.817,0.825],
    [0.108,0.141,0.192,0.236,0.261,0.286,0.317,0.353,0.390,0.426,0.446,0.444,0.423,0.385,0.337,0.283,0.231,0.185,0.146,0.118,0.101,0.090,0.082,0.076,0.074,0.073,0.073,0.074,0.076,0.077,0.076,0.075,0.073,0.072,0.074,0.079],
    [0.189,0.255,0.423,0.660,0.811,0.862,0.877,0.884,0.891,0.896,0.899,0.904,0.907,0.909,0.911,0.910,0.911,0.914,0.913,0.916,0.915,0.916,0.914,0.915,0.918,0.919,0.921,0.923,0.924,0.922,0.922,0.925,0.927,0.930,0.930,0.933],
    [0.171,0.232,0.365,0.507,0.567,0.583,0.588,0.590,0.591,0.590,0.588,0.588,0.589,0.589,0.591,0.590,0.590,0.590,0.589,0.591,0.590,0.590,0.587,0.585,0.583,0.580,0.578,0.576,0.574,0.572,0.571,0.569,0.568,0.568,0.566,0.566],
    [0.144,0.192,0.272,0.331,0.350,0.357,0.361,0.363,0.363,0.361,0.359,0.358,0.358,0.359,0.360,0.360,0.361,0.361,0.360,0.362,0.362,0.361,0.359,0.358,0.355,0.352,0.350,0.348,0.345,0.343,0.340,0.338,0.335,0.334,0.332,0.331],
    [0.105,0.131,0.163,0.180,0.186,0.190,0.193,0.194,0.194,0.192,0.191,0.191,0.191,0.192,0.192,0.192,0.192,0.192,0.192,0.193,0.192,0.192,0.191,0.189,0.188,0.186,0.184,0.182,0.181,0.179,0.178,0.176,0.174,0.173,0.172,0.171],
    [0.068,0.077,0.084,0.087,0.089,0.090,0.092,0.092,0.091,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.089,0.089,0.088,0.087,0.086,0.086,0.085,0.084,0.084,0.083,0.083,0.082,0.081,0.081,0.081],
    [0.031,0.032,0.032,0.033,0.033,0.033,0.033,0.033,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.033],
])
WHITE_IDX = 18
GAMMA = 1.5

_D65_WL  = np.arange(380, 731, 10)
_D65_SPD = np.array([50.00,54.65,82.75,91.49,93.43,86.68,104.86,117.01,
    117.81,114.86,115.92,108.81,109.35,107.80,104.79,107.69,
    104.41,104.33,100.00,96.00,95.12,89.10,90.51,90.32,
    88.18,87.90,83.44,83.84,80.03,80.21,82.28,78.28,
    69.72,71.61,74.35,61.60])

# ── helpers ───────────────────────────────────────────────────────────────────
def interp(wl_src, val_src, wl_dst):
    return np.interp(wl_dst, wl_src, val_src)

def shift_spec(wl, vals, shift):
    if shift == 0: return vals.copy()
    return np.interp(wl, wl - shift, vals, left=vals[0], right=vals[-1])

def load_optics(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["200921再構築"]
    rows = list(ws.iter_rows(min_row=68, max_row=128, values_only=True))
    wb.close()
    def col(i): return np.array([float(r[i]) if r[i] is not None else 1.0 for r in rows])
    return dict(wl=col(1), cf_r=col(3), cf_g=col(6), cf_h=col(7), lens=col(8), ircf=col(9))

def load_ccm(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    vals = [float(r[0]) for r in ws.iter_rows(min_row=20, max_row=28, min_col=2, max_col=2, values_only=True)]
    wb.close()
    return np.array(vals).reshape(3, 3)  # M[i][j]: input i → output j

# ── simulation ────────────────────────────────────────────────────────────────
def simulate_pattern(opt, M_ccm, shift_r, shift_g, shift_b, use_ccm=True):
    wl   = opt["wl"]
    ill  = np.interp(wl, _D65_WL, _D65_SPD)
    ircf = opt["ircf"]
    cf_r = shift_spec(wl, opt["cf_r"], shift_r)
    cf_g = shift_spec(wl, opt["cf_g"], shift_g)
    cf_h = shift_spec(wl, opt["cf_h"], shift_b)
    optical = opt["lens"] * ircf

    raw_integrals = []
    for pi in range(24):
        refl_i = np.interp(wl, WL_MB, REFL[pi])
        inc    = ill * refl_i * optical
        R = np.trapezoid(inc * cf_r, wl)
        G = np.trapezoid(inc * cf_g, wl)
        B = np.trapezoid(inc * cf_h, wl)
        raw_integrals.append((R, G, B))

    # AWB normalization (per-channel white ref)
    wR, wG, wB = raw_integrals[WHITE_IDX]
    results = []
    for R, G, B in raw_integrals:
        Rn, Gn, Bn = R/wR, G/wG, B/wB
        # CCM: [R'G'B'] = [Rn Gn Bn] @ M
        if use_ccm:
            v = np.array([Rn, Gn, Bn]) @ M_ccm
            Rn, Gn, Bn = v[0], v[1], v[2]
        def g8(x): return int(round(np.clip(x, 0, 1)**(1/GAMMA) * 255))
        results.append(dict(
            R_raw=R, G_raw=G, B_raw=B,
            R_norm=R/wR, G_norm=G/wG, B_norm=B/wB,  # pre-CCM
            R_ccm=Rn, G_ccm=Gn, B_ccm=Bn,           # post-CCM (=pre-CCM if CCM off)
            R8=g8(Rn), G8=g8(Gn), B8=g8(Bn),
        ))
    return results, dict(wl=wl, cf_r=cf_r, cf_g=cf_g, cf_h=cf_h,
                         ircf=ircf, lens=opt["lens"])

# ── Excel styling helpers ──────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="0F3460")
HDR_FONT  = Font(color="A8D8EA", bold=True, size=10)
SUB_FILL  = PatternFill("solid", fgColor="16213E")
THIN = Side(style="thin", color="2A2A3E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr(ws, row, col, val, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill or HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER
    return c

def cell(ws, row, col, val, fmt=None, fill=None, color=None):
    c = ws.cell(row=row, column=col, value=val)
    if fmt:   c.number_format = fmt
    if fill:  c.fill = fill
    if color: c.font = Font(color=color, size=10)
    c.border = BORDER
    c.alignment = Alignment(horizontal="right")
    return c

def patch_fill(R8, G8, B8):
    return PatternFill("solid", fgColor=f"{R8:02X}{G8:02X}{B8:02X}")

# ── write per-pattern sheet ───────────────────────────────────────────────────
def write_pattern_sheet(wb, label, results, opt_shifted, M_ccm, use_ccm):
    ws = wb.create_sheet(title=label[:31])
    wl = opt_shifted["wl"]

    # ── Section 1: RGB results ──────────────────────────────────────────────
    r = 1
    ws.cell(r, 1, "■ MacBeth各パッチ RGB値").font = Font(bold=True, color="A8D8EA", size=11)
    ws.cell(r, 1).fill = SUB_FILL
    r += 1
    hdrs = ["No","パッチ名",
            "R_raw(積分)","G_raw(積分)","B_raw(積分)",
            "R_norm(AWB前)","G_norm(AWB前)","B_norm(AWB前)",
            "R_ccm","G_ccm","B_ccm",
            "R8","G8","B8","パッチ色"]
    for ci, h in enumerate(hdrs, 1): hdr(ws, r, ci, h)
    r += 1
    for pi, res in enumerate(results):
        cell(ws, r, 1, pi+1)
        c = ws.cell(r, 2, PATCH_NAMES[pi])
        c.border = BORDER; c.font = Font(size=10)
        cell(ws, r, 3, round(res["R_raw"],4), "0.0000")
        cell(ws, r, 4, round(res["G_raw"],4), "0.0000")
        cell(ws, r, 5, round(res["B_raw"],4), "0.0000")
        cell(ws, r, 6, round(res["R_norm"],4), "0.0000")
        cell(ws, r, 7, round(res["G_norm"],4), "0.0000")
        cell(ws, r, 8, round(res["B_norm"],4), "0.0000")
        cell(ws, r, 9,  round(res["R_ccm"],4), "0.0000", color="FF7F7F" if not use_ccm else None)
        cell(ws, r, 10, round(res["G_ccm"],4), "0.0000", color="7FBF7F" if not use_ccm else None)
        cell(ws, r, 11, round(res["B_ccm"],4), "0.0000", color="7FA0FF" if not use_ccm else None)
        cell(ws, r, 12, res["R8"])
        cell(ws, r, 13, res["G8"])
        cell(ws, r, 14, res["B8"])
        c2 = ws.cell(r, 15, f"RGB({res['R8']},{res['G8']},{res['B8']})")
        c2.fill = patch_fill(res["R8"], res["G8"], res["B8"])
        luma = 0.299*res["R8"] + 0.587*res["G8"] + 0.114*res["B8"]
        c2.font = Font(color="000000" if luma > 128 else "FFFFFF", size=9)
        c2.border = BORDER
        r += 1

    r += 1

    # ── Section 2: CCM matrix ───────────────────────────────────────────────
    ws.cell(r, 1, f"■ CCM {'適用' if use_ccm else '未適用'}  [R′G′B′] = [R G B] × M").font = Font(bold=True, color="A8D8EA", size=11)
    ws.cell(r, 1).fill = SUB_FILL
    r += 1
    for ci, h in enumerate(["(入力＼出力)", "R'", "G'", "B'"], 1): hdr(ws, r, ci, h)
    r += 1
    for i, row_label in enumerate(["R", "G", "B"]):
        cell(ws, r, 1, row_label)
        for j in range(3):
            v = M_ccm[i][j]
            fmt = "+0.0000;-0.0000;0.0000"
            cell(ws, r, j+2, round(v, 4), fmt, color="7EC8E3" if v>=0 else "E08080")
        r += 1

    r += 1

    # ── Section 3: 分光特性 ─────────────────────────────────────────────────
    ws.cell(r, 1, "■ 分光特性（シフト後）").font = Font(bold=True, color="A8D8EA", size=11)
    ws.cell(r, 1).fill = SUB_FILL
    r += 1
    spec_hdrs = ["波長(nm)", "レンズ", "IRCF", "CF-R", "CF-G", "CF-B", "レンズ×IRCF"]
    for ci, h in enumerate(spec_hdrs, 1): hdr(ws, r, ci, h)
    r += 1
    for wi, w in enumerate(wl):
        cell(ws, r, 1, int(w))
        cell(ws, r, 2, round(float(opt_shifted["lens"][wi]),5), "0.00000")
        cell(ws, r, 3, round(float(opt_shifted["ircf"][wi]),5), "0.00000")
        cell(ws, r, 4, round(float(opt_shifted["cf_r"][wi]),5), "0.00000", color="FF9090")
        cell(ws, r, 5, round(float(opt_shifted["cf_g"][wi]),5), "0.00000", color="90EE90")
        cell(ws, r, 6, round(float(opt_shifted["cf_h"][wi]),5), "0.00000", color="90B0FF")
        cell(ws, r, 7, round(float(opt_shifted["lens"][wi]*opt_shifted["ircf"][wi]),5), "0.00000")
        r += 1

    r += 1

    # ── Section 4: マクベス反射率 ───────────────────────────────────────────
    ws.cell(r, 1, "■ マクベス分光反射率 (BabelColor Avg, 380-730nm 10nm)").font = Font(bold=True, color="A8D8EA", size=11)
    ws.cell(r, 1).fill = SUB_FILL
    r += 1
    ref_hdrs = ["波長(nm)"] + [n.replace(f"{i+1} ","") for i,n in enumerate(PATCH_NAMES)]
    for ci, h in enumerate(ref_hdrs, 1): hdr(ws, r, ci, h)
    r += 1
    for wi, w in enumerate(WL_MB):
        cell(ws, r, 1, int(w))
        for pi in range(24):
            cell(ws, r, pi+2, round(float(REFL[pi][wi]),4), "0.0000")
        r += 1

    # column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    for c in range(3, 20): ws.column_dimensions[get_column_letter(c)].width = 11

# ── write summary sheet ───────────────────────────────────────────────────────
def write_summary_sheet(wb, patterns):
    ws = wb.create_sheet(title="Summary", index=0)
    r = 1
    ws.cell(r, 1, "MacBeth Simulation 一括結果サマリ (D65, gamma=1.5, CCMあり)").font = Font(bold=True, color="A8D8EA", size=12)
    ws.cell(r, 1).fill = SUB_FILL
    r += 2

    # header: No / Name / then per-pattern R8,G8,B8
    base_cols = 2
    hdr(ws, r, 1, "No")
    hdr(ws, r, 2, "パッチ名")
    for pi_p, (label, _) in enumerate(patterns):
        col = base_cols + 1 + pi_p * 3
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+2)
        c = ws.cell(r, col, label)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    r += 1
    hdr(ws, r, 1, "No"); hdr(ws, r, 2, "パッチ名")
    for pi_p in range(len(patterns)):
        col = base_cols + 1 + pi_p * 3
        hdr(ws, r, col,   "R8")
        hdr(ws, r, col+1, "G8")
        hdr(ws, r, col+2, "B8")
    r += 1

    for pi in range(24):
        cell(ws, r, 1, pi+1)
        c = ws.cell(r, 2, PATCH_NAMES[pi])
        c.border = BORDER; c.font = Font(size=10)
        for pi_p, (_, results) in enumerate(patterns):
            col = base_cols + 1 + pi_p * 3
            res = results[pi]
            cell(ws, r, col,   res["R8"])
            cell(ws, r, col+1, res["G8"])
            cell(ws, r, col+2, res["B8"])
            # color the B8 cell as patch background
            c2 = ws.cell(r, col+2)
            c2.fill = patch_fill(res["R8"], res["G8"], res["B8"])
        r += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    for pi_p in range(len(patterns)):
        for j in range(3):
            ws.column_dimensions[get_column_letter(base_cols + 1 + pi_p*3 + j)].width = 7

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel",  default="201006_車両WSカメラ影響定量化まとめ２.xlsx")
    p.add_argument("--ccm",    default="ccm_coef.xlsx")
    p.add_argument("--out",    default="batch_result.xlsx")
    p.add_argument("--shift",  type=int, default=10,
                   help="CF shift amount in nm (default: 10)")
    args = p.parse_args()

    print(f"Loading optics: {args.excel}")
    opt = load_optics(args.excel)
    print(f"Loading CCM: {args.ccm}")
    M = load_ccm(args.ccm)

    # 8 patterns: R/G/B each 0 or +shift
    S = args.shift
    patterns_raw = list(itertools.product([0, S], repeat=3))  # (sR, sG, sB)

    print(f"Running {len(patterns_raw)} patterns (CF shift ±{S}nm)...")
    pattern_results = []
    for sR, sG, sB in patterns_raw:
        label = f"R{sR:+d}G{sG:+d}B{sB:+d}"
        results, opt_shifted = simulate_pattern(opt, M, sR, sG, sB, use_ccm=True)
        pattern_results.append((label, results, opt_shifted))
        print(f"  {label}")

    print(f"Writing {args.out} ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Summary first
    write_summary_sheet(wb, [(label, res) for label, res, _ in pattern_results])

    # Per-pattern sheets
    for label, results, opt_shifted in pattern_results:
        write_pattern_sheet(wb, label, results, opt_shifted, M, use_ccm=True)

    wb.save(args.out)
    print(f"Done: {args.out}")
    print(f"  Sheets: Summary + {len(pattern_results)} pattern sheets")

if __name__ == "__main__":
    main()
