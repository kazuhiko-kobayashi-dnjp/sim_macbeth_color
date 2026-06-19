#!/usr/bin/env python3
"""
MacBeth ColorChecker RGB simulation
Illuminant × Reflectance × Lens × IRCF × ColorFilter → R, G, H integrals

Usage:
  python sim_macbeth.py --excel optics.xlsx
  python sim_macbeth.py --excel optics.xlsx --shift-cf 10 --shift-ircf -10
  python sim_macbeth.py --excel optics.xlsx --illuminant A --tone-curve 2.2
"""
import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

# ─── MacBeth wavelength axis: 10 nm steps, 380–730 nm ─────────────────────────
WL_MB = np.arange(380, 731, 10)  # 36 points

# ─── BabelColor Avg reflectance, shape (24, 36) ────────────────────────────────
PATCH_NAMES = [
    "1 dark skin",    "2 light skin",   "3 blue sky",      "4 foliage",
    "5 blue flower",  "6 bluish green", "7 orange",        "8 purplish blue",
    "9 moderate red", "10 purple",      "11 yellow green", "12 orange yellow",
    "13 blue",        "14 green",       "15 red",          "16 yellow",
    "17 magenta",     "18 cyan",        "19 white 9.5",    "20 neutral 8",
    "21 neutral 6.5", "22 neutral 5",   "23 neutral 3.5",  "24 black 2",
]

REFL = np.array([
    [0.055,0.058,0.061,0.062,0.062,0.062,0.062,0.062,0.062,0.062,0.062,0.063,0.065,0.070,0.076,0.079,0.081,0.084,0.091,0.103,0.119,0.134,0.143,0.147,0.151,0.158,0.168,0.179,0.188,0.190,0.186,0.181,0.182,0.187,0.196,0.209],
    [0.117,0.143,0.175,0.191,0.196,0.199,0.204,0.213,0.228,0.251,0.280,0.309,0.329,0.333,0.315,0.286,0.273,0.276,0.277,0.289,0.339,0.420,0.488,0.525,0.546,0.562,0.578,0.595,0.612,0.625,0.638,0.656,0.678,0.700,0.717,0.734],
    [0.130,0.177,0.251,0.306,0.324,0.330,0.333,0.331,0.323,0.311,0.298,0.285,0.269,0.250,0.231,0.214,0.199,0.185,0.169,0.157,0.149,0.145,0.142,0.141,0.141,0.141,0.143,0.147,0.152,0.154,0.150,0.144,0.136,0.132,0.135,0.147],
    [0.051,0.054,0.056,0.057,0.058,0.059,0.060,0.061,0.062,0.063,0.065,0.067,0.075,0.101,0.145,0.178,0.184,0.170,0.149,0.133,0.122,0.115,0.109,0.105,0.104,0.106,0.109,0.112,0.114,0.114,0.112,0.112,0.115,0.120,0.125,0.130],
    [0.144,0.198,0.294,0.375,0.408,0.421,0.426,0.426,0.419,0.403,0.379,0.346,0.311,0.281,0.254,0.229,0.214,0.208,0.202,0.194,0.193,0.200,0.214,0.230,0.241,0.254,0.279,0.313,0.348,0.366,0.366,0.359,0.358,0.365,0.377,0.398],
    [0.136,0.179,0.247,0.297,0.320,0.337,0.355,0.381,0.419,0.466,0.510,0.546,0.567,0.574,0.569,0.551,0.524,0.488,0.445,0.400,0.350,0.299,0.252,0.221,0.204,0.196,0.191,0.188,0.191,0.199,0.212,0.223,0.232,0.233,0.229,0.229],
    [0.054,0.054,0.053,0.054,0.054,0.055,0.055,0.055,0.056,0.057,0.058,0.061,0.068,0.089,0.125,0.154,0.174,0.199,0.248,0.335,0.444,0.538,0.587,0.595,0.591,0.587,0.584,0.584,0.590,0.603,0.620,0.639,0.655,0.663,0.663,0.667],
    [0.122,0.164,0.229,0.286,0.327,0.361,0.388,0.400,0.392,0.362,0.316,0.260,0.209,0.168,0.138,0.117,0.104,0.096,0.090,0.086,0.084,0.084,0.084,0.084,0.084,0.085,0.090,0.098,0.109,0.123,0.143,0.169,0.205,0.244,0.287,0.332],
    [0.096,0.115,0.131,0.135,0.133,0.132,0.130,0.128,0.125,0.120,0.115,0.110,0.105,0.100,0.095,0.093,0.092,0.093,0.096,0.108,0.156,0.265,0.399,0.500,0.556,0.579,0.588,0.591,0.593,0.594,0.598,0.602,0.607,0.609,0.609,0.610],
    [0.092,0.116,0.146,0.169,0.178,0.173,0.158,0.139,0.119,0.101,0.087,0.075,0.066,0.060,0.056,0.053,0.051,0.051,0.052,0.052,0.051,0.052,0.058,0.073,0.096,0.119,0.141,0.166,0.194,0.227,0.265,0.309,0.355,0.396,0.436,0.478],
    [0.061,0.061,0.062,0.063,0.064,0.066,0.069,0.075,0.085,0.105,0.139,0.192,0.271,0.376,0.476,0.531,0.549,0.546,0.528,0.504,0.471,0.428,0.381,0.347,0.327,0.318,0.312,0.310,0.314,0.327,0.345,0.363,0.376,0.381,0.378,0.379],
    [0.063,0.063,0.063,0.064,0.064,0.064,0.065,0.066,0.067,0.068,0.071,0.076,0.087,0.125,0.206,0.305,0.383,0.431,0.469,0.518,0.568,0.607,0.628,0.637,0.640,0.642,0.645,0.648,0.651,0.653,0.657,0.664,0.673,0.680,0.684,0.688],
    [0.066,0.079,0.102,0.146,0.200,0.244,0.282,0.309,0.308,0.278,0.231,0.178,0.130,0.094,0.070,0.054,0.046,0.042,0.039,0.038,0.038,0.038,0.038,0.039,0.039,0.040,0.041,0.042,0.044,0.045,0.046,0.046,0.048,0.052,0.057,0.065],
    [0.052,0.053,0.054,0.055,0.057,0.059,0.061,0.066,0.075,0.093,0.125,0.178,0.246,0.307,0.337,0.334,0.317,0.293,0.262,0.230,0.198,0.165,0.135,0.115,0.104,0.098,0.094,0.092,0.093,0.097,0.102,0.108,0.113,0.115,0.114,0.114],
    [0.050,0.049,0.048,0.047,0.047,0.047,0.047,0.047,0.046,0.045,0.044,0.044,0.045,0.046,0.047,0.048,0.049,0.050,0.054,0.060,0.072,0.104,0.178,0.312,0.467,0.581,0.644,0.675,0.690,0.698,0.706,0.715,0.724,0.730,0.734,0.738],
    [0.058,0.054,0.052,0.052,0.053,0.054,0.056,0.059,0.067,0.081,0.107,0.152,0.225,0.336,0.462,0.559,0.616,0.650,0.672,0.694,0.710,0.723,0.731,0.739,0.746,0.752,0.758,0.764,0.769,0.771,0.776,0.782,0.790,0.796,0.799,0.804],
    [0.145,0.195,0.283,0.346,0.362,0.354,0.334,0.306,0.276,0.248,0.218,0.190,0.168,0.149,0.127,0.107,0.100,0.102,0.104,0.109,0.137,0.200,0.290,0.400,0.516,0.615,0.687,0.732,0.760,0.774,0.783,0.793,0.803,0.812,0.817,0.825],
    [0.108,0.141,0.192,0.236,0.261,0.286,0.317,0.353,0.390,0.426,0.446,0.444,0.423,0.385,0.337,0.283,0.231,0.185,0.146,0.118,0.101,0.090,0.082,0.076,0.074,0.073,0.073,0.074,0.076,0.077,0.076,0.075,0.073,0.072,0.074,0.079],
    [0.189,0.255,0.423,0.660,0.811,0.862,0.877,0.884,0.891,0.896,0.899,0.904,0.907,0.909,0.911,0.910,0.911,0.914,0.913,0.916,0.915,0.916,0.914,0.915,0.918,0.919,0.921,0.923,0.924,0.922,0.922,0.925,0.927,0.930,0.930,0.933],
    [0.171,0.232,0.365,0.507,0.567,0.583,0.588,0.590,0.591,0.590,0.588,0.588,0.589,0.589,0.591,0.590,0.590,0.590,0.589,0.591,0.590,0.590,0.587,0.585,0.583,0.580,0.578,0.576,0.574,0.572,0.571,0.569,0.568,0.568,0.566,0.566],
    [0.144,0.192,0.272,0.331,0.350,0.357,0.361,0.363,0.363,0.361,0.359,0.358,0.358,0.359,0.360,0.360,0.361,0.361,0.360,0.362,0.362,0.361,0.359,0.358,0.355,0.352,0.350,0.348,0.345,0.343,0.340,0.338,0.335,0.334,0.332,0.331],
    [0.105,0.131,0.163,0.180,0.186,0.190,0.193,0.194,0.194,0.192,0.191,0.191,0.191,0.192,0.192,0.192,0.192,0.192,0.192,0.193,0.192,0.192,0.191,0.189,0.188,0.186,0.184,0.182,0.181,0.179,0.178,0.176,0.174,0.173,0.172,0.171],
    [0.068,0.077,0.084,0.087,0.089,0.090,0.092,0.092,0.091,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.090,0.089,0.089,0.088,0.087,0.086,0.086,0.085,0.084,0.084,0.083,0.083,0.082,0.081,0.081,0.081],
    [0.031,0.032,0.032,0.033,0.033,0.033,0.033,0.033,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.032,0.033],
])  # shape (24, 36)

WHITE_IDX = 18  # "19 white 9.5" (0-indexed)

# ─── CIE illuminants ──────────────────────────────────────────────────────────

def illuminant_a(wl: np.ndarray) -> np.ndarray:
    """CIE Standard Illuminant A (2856 K), normalized to 100 at 560 nm."""
    c2 = 1.4388e7  # nm·K
    T = 2856.0
    s = (560.0 / wl) ** 5 * (np.exp(c2 / (T * 560.0)) - 1.0) / (np.exp(c2 / (T * wl)) - 1.0)
    return s * 100.0 / np.interp(560.0, wl, s)


# CIE D65 tabulated at 10 nm, 380–730 nm (CIE Pub. 15, Table 1)
_D65_WL = np.arange(380, 731, 10)
_D65_SPD = np.array([
     50.00,  54.65,  82.75,  91.49,  93.43,  86.68, 104.86, 117.01,
    117.81, 114.86, 115.92, 108.81, 109.35, 107.80, 104.79, 107.69,
    104.41, 104.33, 100.00,  96.00,  95.12,  89.10,  90.51,  90.32,
     88.18,  87.90,  83.44,  83.84,  80.03,  80.21,  82.28,  78.28,
     69.72,  71.61,  74.35,  61.60,
])


def illuminant_d65(wl: np.ndarray) -> np.ndarray:
    return np.interp(wl, _D65_WL, _D65_SPD)


ILLUMINANTS = {"A": illuminant_a, "D65": illuminant_d65}

# ─── Wavelength shift ─────────────────────────────────────────────────────────

def shift_spectrum(wl: np.ndarray, spec: np.ndarray, shift_nm: float) -> np.ndarray:
    """Shift spectrum by +shift_nm (positive = redshift). Edges clamped."""
    if shift_nm == 0:
        return spec
    return np.interp(wl - shift_nm, wl, spec, left=spec[0], right=spec[-1])

# ─── Load optics from Excel ───────────────────────────────────────────────────

EXCEL_SHEET = "200921再構築"
EXCEL_ROW_START = 68   # 1-indexed (inclusive)
EXCEL_ROW_END   = 128  # 1-indexed (inclusive)
# col indices (0-indexed within row tuple): B=1, D=3, E=4, F=5, I=8, J=9


def load_optics(excel_path: str, sheet: str = EXCEL_SHEET) -> dict:
    """
    Read wavelength + CF_R/G/H + Lens + IRCF from Excel.
    Returns dict: wl, cf_r, cf_g, cf_h, lens, ircf  (all float numpy arrays)
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
    except KeyError:
        sheets = wb.sheetnames
        wb.close()
        raise ValueError(f"Sheet '{sheet}' not found. Available: {sheets}")

    rows = list(ws.iter_rows(
        min_row=EXCEL_ROW_START, max_row=EXCEL_ROW_END, values_only=True
    ))
    wb.close()

    def col(idx):
        vals = []
        for r in rows:
            v = r[idx] if idx < len(r) else None
            vals.append(float(v) if v is not None else np.nan)
        return np.array(vals, dtype=float)

    wl    = col(1)   # B
    cf_r  = col(3)   # D  CF Red
    cf_g  = col(6)   # G  CF Green (avg of Gr/Gb)
    cf_h  = col(7)   # H  CF Blue
    lens  = col(8)   # I  Lens
    ircf  = col(9)   # J  IRCF

    # Replace NaN with 1 (pass-through) and warn
    for name, arr in [("cf_r", cf_r), ("cf_g", cf_g), ("cf_h", cf_h),
                      ("lens", lens), ("ircf", ircf)]:
        n = np.isnan(arr).sum()
        if n:
            print(f"  WARNING: {n} NaN in '{name}', replaced with 1.0")
            arr[np.isnan(arr)] = 1.0

    return dict(wl=wl, cf_r=cf_r, cf_g=cf_g, cf_h=cf_h, lens=lens, ircf=ircf)


def unity_optics(wl_start=400, wl_end=700, wl_step=5) -> dict:
    """Fallback: all transmittances = 1 on a 5 nm grid 400–700 nm."""
    wl = np.arange(wl_start, wl_end + 1, wl_step, dtype=float)
    ones = np.ones(len(wl))
    return dict(wl=wl, cf_r=ones.copy(), cf_g=ones.copy(), cf_h=ones.copy(),
                lens=ones.copy(), ircf=ones.copy())

# ─── Simulation ───────────────────────────────────────────────────────────────

def simulate(optics: dict, ill_func,
             shift_cf: float = 0.0,
             shift_ircf: float = 0.0,
             gamma: float = 0.0) -> pd.DataFrame:
    """
    Compute R, G, H integrals for all 24 MacBeth patches.

    optics     – dict from load_optics() or unity_optics()
    ill_func   – callable(wl) -> SPD array
    shift_cf   – CF wavelength shift in nm (+10 = redshift)
    shift_ircf – IRCF cutoff shift in nm
    gamma      – if > 0, apply x^(1/gamma) tone curve and output 8-bit values

    Returns DataFrame with columns:
      No, Name, R_raw, G_raw, H_raw,
      R_norm, G_norm, H_norm,   (relative to white patch)
      [R8, G8, H8]              (8-bit, only if gamma > 0)
    """
    wl = optics["wl"]

    ill   = ill_func(wl)
    lens  = optics["lens"]
    ircf  = shift_spectrum(wl, optics["ircf"], shift_ircf)
    cf_r  = shift_spectrum(wl, optics["cf_r"], shift_cf)
    cf_g  = shift_spectrum(wl, optics["cf_g"], shift_cf)
    cf_h  = shift_spectrum(wl, optics["cf_h"], shift_cf)

    optical = lens * ircf  # common optical path

    records = []
    for i, name in enumerate(PATCH_NAMES):
        refl = np.interp(wl, WL_MB, REFL[i])
        incident = ill * refl * optical
        r = np.trapezoid(incident * cf_r, wl)
        g = np.trapezoid(incident * cf_g, wl)
        h = np.trapezoid(incident * cf_h, wl)
        records.append({"No": i + 1, "Name": name, "R_raw": r, "G_raw": g, "H_raw": h})

    df = pd.DataFrame(records)

    white = df.loc[WHITE_IDX, ["R_raw", "G_raw", "H_raw"]]
    df["R_norm"] = df["R_raw"] / white["R_raw"]
    df["G_norm"] = df["G_raw"] / white["G_raw"]
    df["H_norm"] = df["H_raw"] / white["H_raw"]

    if gamma > 0:
        clip = lambda x: np.clip(x, 0.0, 1.0)
        df["R8"] = (clip(df["R_norm"]) ** (1.0 / gamma) * 255).round().astype(int)
        df["G8"] = (clip(df["G_norm"]) ** (1.0 / gamma) * 255).round().astype(int)
        df["H8"] = (clip(df["H_norm"]) ** (1.0 / gamma) * 255).round().astype(int)

    return df

# ─── HTML output ─────────────────────────────────────────────────────────────

def _cell_color(r, g, b):
    """sRGB hex from 0–1 floats."""
    r8 = int(np.clip(r, 0, 1) * 255)
    g8 = int(np.clip(g, 0, 1) * 255)
    b8 = int(np.clip(b, 0, 1) * 255)
    luma = 0.299 * r8 + 0.587 * g8 + 0.114 * b8
    fg = "#000" if luma > 128 else "#fff"
    return f"#{r8:02x}{g8:02x}{b8:02x}", fg


def build_html(results: dict) -> str:
    """
    results: {label: DataFrame}  – one entry per illuminant / scenario
    Renders a MacBeth-style 4×6 color grid + raw value table for each scenario.
    H channel is mapped to Blue in the preview (adjust if different sensor).
    """
    css = """
<style>
body { font-family: sans-serif; margin: 20px; }
h2 { margin-top: 40px; }
.chart { display: grid; grid-template-columns: repeat(6, 120px);
         gap: 4px; margin-bottom: 20px; }
.patch { width: 120px; height: 70px; display: flex; flex-direction: column;
         align-items: center; justify-content: center;
         font-size: 10px; border-radius: 4px; }
.patch .pname { font-weight: bold; }
table { border-collapse: collapse; font-size: 12px; margin-top: 10px; }
th, td { border: 1px solid #ccc; padding: 4px 8px; text-align: right; }
th { background: #eee; text-align: center; }
td.name { text-align: left; }
</style>
"""
    body_parts = [f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
                  f"<title>MacBeth Simulation</title>{css}</head><body>"
                  f"<h1>MacBeth ColorChecker RGB Simulation</h1>"]

    for label, df in results.items():
        body_parts.append(f"<h2>{label}</h2>")

        # Color grid (4 rows × 6 cols)
        body_parts.append('<div class="chart">')
        for _, row in df.iterrows():
            bg, fg = _cell_color(row["R_norm"], row["G_norm"], row["H_norm"])
            name_short = row["Name"].split(" ", 1)[1] if " " in row["Name"] else row["Name"]
            r_str = f"{row['R_norm']:.3f}"
            g_str = f"{row['G_norm']:.3f}"
            h_str = f"{row['H_norm']:.3f}"
            body_parts.append(
                f'<div class="patch" style="background:{bg};color:{fg}">'
                f'<span class="pname">{name_short}</span>'
                f'<span>R={r_str} G={g_str}</span>'
                f'<span>H(B)={h_str}</span></div>'
            )
        body_parts.append("</div>")

        # Numeric table
        has8 = "R8" in df.columns
        cols8 = ["R8", "G8", "H8"] if has8 else []
        headers = ["No", "Name", "R_raw", "G_raw", "H_raw",
                   "R_norm", "G_norm", "H_norm"] + cols8
        body_parts.append("<table><tr>" +
                          "".join(f"<th>{h}</th>" for h in headers) + "</tr>")
        for _, row in df.iterrows():
            body_parts.append("<tr>")
            body_parts.append(f"<td>{int(row['No'])}</td>")
            body_parts.append(f'<td class="name">{row["Name"]}</td>')
            for c in ["R_raw", "G_raw", "H_raw"]:
                body_parts.append(f"<td>{row[c]:.4f}</td>")
            for c in ["R_norm", "G_norm", "H_norm"]:
                body_parts.append(f"<td>{row[c]:.4f}</td>")
            for c in cols8:
                body_parts.append(f"<td>{int(row[c])}</td>")
            body_parts.append("</tr>")
        body_parts.append("</table>")

    body_parts.append("</body></html>")
    return "\n".join(body_parts)

# ─── Excel output ────────────────────────────────────────────────────────────

def save_excel(results: dict, path: Path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Summary sheet: all scenarios side-by-side
        summary_frames = []
        for label, df in results.items():
            sub = df[["No", "Name"]].copy()
            has8 = "R8" in df.columns
            for c in ["R_norm", "G_norm", "H_norm"] + (["R8", "G8", "H8"] if has8 else []):
                sub[f"{label}_{c}"] = df[c]
            summary_frames.append(sub)

        merged = summary_frames[0]
        for sf in summary_frames[1:]:
            extra_cols = [c for c in sf.columns if c not in ("No", "Name")]
            merged = merged.merge(sf[["No"] + extra_cols], on="No")
        merged.to_excel(writer, sheet_name="Summary", index=False)

        # Per-scenario sheets
        for label, df in results.items():
            sheet = label[:31]  # Excel sheet name limit
            df.to_excel(writer, sheet_name=sheet, index=False)

# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="MacBeth RGB simulation")
    p.add_argument("--excel", help="Path to optics Excel file")
    p.add_argument("--sheet", default=EXCEL_SHEET, help="Sheet name in Excel")
    p.add_argument("--illuminant", choices=["A", "D65", "both"], default="both")
    p.add_argument("--shift-cf",   type=float, default=0.0,
                   help="CF wavelength shift in nm (+10 = redshift)")
    p.add_argument("--shift-ircf", type=float, default=0.0,
                   help="IRCF cutoff shift in nm")
    p.add_argument("--gamma", type=float, default=0.0,
                   help="Tone curve gamma (e.g. 2.2). 0 = no 8-bit output")
    p.add_argument("--out", default="macbeth_sim", help="Output base name (no extension)")
    args = p.parse_args()

    # Load optics
    if args.excel:
        print(f"Loading optics from: {args.excel}")
        optics = load_optics(args.excel, sheet=args.sheet)
        wl = optics["wl"]
        print(f"  Wavelength range: {wl[0]:.0f}–{wl[-1]:.0f} nm, "
              f"{len(wl)} points, step {wl[1]-wl[0]:.0f} nm")
    else:
        print("WARNING: --excel not specified. Using unity transmittance "
              "(flat 1.0 for lens/IRCF/CF on 400–700 nm, 5 nm grid).")
        optics = unity_optics()

    # Illuminants to run
    ill_names = []
    if args.illuminant in ("A", "both"):
        ill_names.append("A")
    if args.illuminant in ("D65", "both"):
        ill_names.append("D65")

    # Build scenario label
    suffix = ""
    if args.shift_cf:
        suffix += f"_CF{args.shift_cf:+.0f}nm"
    if args.shift_ircf:
        suffix += f"_IRCF{args.shift_ircf:+.0f}nm"

    results = {}
    for name in ill_names:
        label = f"{name}{suffix}"
        print(f"Simulating: {label} ...")
        df = simulate(optics, ILLUMINANTS[name],
                      shift_cf=args.shift_cf,
                      shift_ircf=args.shift_ircf,
                      gamma=args.gamma)
        results[label] = df

    # Save outputs
    out_base = Path(args.out)
    xlsx_path = out_base.with_suffix(".xlsx")
    html_path = out_base.with_suffix(".html")

    save_excel(results, xlsx_path)
    print(f"Saved: {xlsx_path}")

    html_path.write_text(build_html(results), encoding="utf-8")
    print(f"Saved: {html_path}")

    # Print summary to console
    for label, df in results.items():
        print(f"\n── {label} ──")
        cols = ["No", "Name", "R_norm", "G_norm", "H_norm"]
        if "R8" in df.columns:
            cols += ["R8", "G8", "H8"]
        print(df[cols].to_string(index=False, float_format="{:.4f}".format))


if __name__ == "__main__":
    main()
